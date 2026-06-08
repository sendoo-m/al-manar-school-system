# students/services/export_service.py
import csv
import io
import json
from datetime import datetime
from decimal import Decimal

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class StudentExportService:
    """خدمة تصدير بيانات الطلاب بعد توسيع ملف الطالب"""

    HEADERS = [
        'اسم الطالب',
        'نوع الطالب',
        'الرقم القومي',
        'رقم جواز السفر',
        'الجنسية',
        'الديانة',
        'العمر',
        'النوع',
        'تاريخ الميلاد',
        'رقم الهاتف',
        'العنوان',

        'العام الدراسي',
        'المرحلة التعليمية',
        'الصف الدراسي',
        'حالة القيد',
        'محول من مدرسة',
        'محول إلى مدرسة',

        'طالب دمج',
        'نوع الإعاقة',
        'إعفاء من العربي',
        'إعفاء من الإنجليزي',
        'إعفاء من الفرنسي',
        'إعفاءات أخرى',

        'من أبناء العاملين',
        'اسم الموظف',
        'وظيفة الموظف داخل المدرسة',

        'اسم ولي الأمر',
        'هاتف ولي الأمر',
        'بريد ولي الأمر',
        'وظيفة الأب',
        'صاحب الولاية التعليمية',
        'اسم صاحب الولاية التعليمية',
        'هاتف صاحب الولاية التعليمية',

        'إجمالي المصروفات',
        'إجمالي المدفوعات',
        'إجمالي المستحقات',
        'الحالة المالية',

        'تاريخ التسجيل',
        'آخر تحديث',
        'نشط',
    ]

    @classmethod
    def export(cls, queryset, export_format='csv'):
        export_format = (export_format or 'csv').lower()

        if export_format in ['excel', 'xlsx']:
            return cls.export_excel(queryset)

        if export_format == 'json':
            return cls.export_json(queryset)

        return cls.export_csv(queryset)

    @classmethod
    def get_student_row(cls, student):
        grade_level = getattr(student, 'grade_level', None)
        education_level = getattr(grade_level, 'education_level', None) if grade_level else None

        return [
            student.name or '',
            student.get_student_type_display() if hasattr(student, 'get_student_type_display') else getattr(student, 'student_type', ''),
            student.national_number or '',
            getattr(student, 'passport_number', '') or '',
            getattr(student, 'nationality', '') or '',
            student.get_religion_display() if hasattr(student, 'get_religion_display') else getattr(student, 'religion', ''),
            student.age or '',
            student.get_gender_display() if hasattr(student, 'get_gender_display') else getattr(student, 'gender', ''),
            student.date_of_birth.strftime('%Y-%m-%d') if getattr(student, 'date_of_birth', None) else '',
            student.phone_number or '',
            student.address or '',

            student.academic_year.name if getattr(student, 'academic_year', None) else '',
            education_level.name if education_level else '',
            grade_level.name if grade_level else '',
            student.get_enrollment_status_display() if hasattr(student, 'get_enrollment_status_display') else getattr(student, 'enrollment_status', ''),
            getattr(student, 'transferred_from_school', '') or '',
            getattr(student, 'transferred_to_school', '') or '',

            'نعم' if getattr(student, 'is_integration_student', False) else 'لا',
            getattr(student, 'disability_type', '') or '',
            'نعم' if getattr(student, 'exempt_from_arabic', False) else 'لا',
            'نعم' if getattr(student, 'exempt_from_english', False) else 'لا',
            'نعم' if getattr(student, 'exempt_from_french', False) else 'لا',
            getattr(student, 'other_subject_exemptions', '') or '',

            'نعم' if getattr(student, 'is_staff_child', False) else 'لا',
            getattr(student, 'staff_parent_name', '') or '',
            getattr(student, 'staff_parent_job', '') or '',

            student.parent_name or '',
            student.parent_phone or '',
            student.parent_email or '',
            getattr(student, 'father_job', '') or '',
            student.get_educational_guardian_display() if hasattr(student, 'get_educational_guardian_display') else getattr(student, 'educational_guardian', ''),
            getattr(student, 'educational_guardian_name', '') or '',
            getattr(student, 'educational_guardian_phone', '') or '',

            float(student.total_fees or 0),
            float(student.total_payments or 0),
            float(student.total_owed or 0),
            student.get_financial_status() if hasattr(student, 'get_financial_status') else '',

            timezone.localtime(student.created_at).strftime('%Y-%m-%d %H:%M') if getattr(student, 'created_at', None) else '',
            timezone.localtime(student.updated_at).strftime('%Y-%m-%d %H:%M') if getattr(student, 'updated_at', None) else '',
            'نعم' if getattr(student, 'is_active', False) else 'لا',
        ]

    @classmethod
    def get_student_dict(cls, student):
        grade_level = getattr(student, 'grade_level', None)
        education_level = getattr(grade_level, 'education_level', None) if grade_level else None

        return {
            'id': student.id,
            'name': student.name or '',
            'student_type': getattr(student, 'student_type', ''),
            'student_type_display': student.get_student_type_display() if hasattr(student, 'get_student_type_display') else '',
            'national_number': student.national_number or '',
            'passport_number': getattr(student, 'passport_number', '') or '',
            'nationality': getattr(student, 'nationality', '') or '',
            'religion': getattr(student, 'religion', '') or '',
            'religion_display': student.get_religion_display() if hasattr(student, 'get_religion_display') else '',
            'age': student.age,
            'gender': student.gender or '',
            'gender_display': student.get_gender_display() if hasattr(student, 'get_gender_display') else '',
            'date_of_birth': student.date_of_birth.isoformat() if getattr(student, 'date_of_birth', None) else '',
            'phone_number': student.phone_number or '',
            'address': student.address or '',

            'academic_year': student.academic_year.name if getattr(student, 'academic_year', None) else '',
            'education_level': education_level.name if education_level else '',
            'grade_level': grade_level.name if grade_level else '',
            'enrollment_status': getattr(student, 'enrollment_status', ''),
            'enrollment_status_display': student.get_enrollment_status_display() if hasattr(student, 'get_enrollment_status_display') else '',
            'transferred_from_school': getattr(student, 'transferred_from_school', '') or '',
            'transferred_to_school': getattr(student, 'transferred_to_school', '') or '',

            'is_integration_student': bool(getattr(student, 'is_integration_student', False)),
            'disability_type': getattr(student, 'disability_type', '') or '',
            'exempt_from_arabic': bool(getattr(student, 'exempt_from_arabic', False)),
            'exempt_from_english': bool(getattr(student, 'exempt_from_english', False)),
            'exempt_from_french': bool(getattr(student, 'exempt_from_french', False)),
            'other_subject_exemptions': getattr(student, 'other_subject_exemptions', '') or '',

            'is_staff_child': bool(getattr(student, 'is_staff_child', False)),
            'staff_parent_name': getattr(student, 'staff_parent_name', '') or '',
            'staff_parent_job': getattr(student, 'staff_parent_job', '') or '',

            'parent_name': student.parent_name or '',
            'parent_phone': student.parent_phone or '',
            'parent_email': student.parent_email or '',
            'father_job': getattr(student, 'father_job', '') or '',
            'educational_guardian': getattr(student, 'educational_guardian', ''),
            'educational_guardian_display': student.get_educational_guardian_display() if hasattr(student, 'get_educational_guardian_display') else '',
            'educational_guardian_name': getattr(student, 'educational_guardian_name', '') or '',
            'educational_guardian_phone': getattr(student, 'educational_guardian_phone', '') or '',

            'total_fees': float(student.total_fees or 0),
            'total_payments': float(student.total_payments or 0),
            'total_owed': float(student.total_owed or 0),
            'financial_status': student.get_financial_status() if hasattr(student, 'get_financial_status') else '',

            'created_at': student.created_at.isoformat() if getattr(student, 'created_at', None) else '',
            'updated_at': student.updated_at.isoformat() if getattr(student, 'updated_at', None) else '',
            'is_active': bool(getattr(student, 'is_active', False)),
        }

    @classmethod
    def export_csv(cls, queryset):
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow(cls.HEADERS)

        for student in queryset:
            writer.writerow(cls.get_student_row(student))

        return response

    @classmethod
    def export_excel(cls, queryset):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'بيانات الطلاب'
        worksheet.sheet_view.rightToLeft = True

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num, header in enumerate(cls.HEADERS, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for row_num, student in enumerate(queryset, 2):
            row = cls.get_student_row(student)

            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                if cls.HEADERS[col_num - 1] in ['إجمالي المصروفات', 'إجمالي المدفوعات', 'إجمالي المستحقات']:
                    cell.number_format = '#,##0.00'

        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions

        for col_num, header in enumerate(cls.HEADERS, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(header))

            for row_num in range(2, worksheet.max_row + 1):
                value = worksheet.cell(row=row_num, column=col_num).value
                max_length = max(max_length, len(str(value)) if value is not None else 0)

            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 35)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'

        return response

    @classmethod
    def export_json(cls, queryset):
        data = [cls.get_student_dict(student) for student in queryset]

        response = HttpResponse(
            json.dumps(
                {
                    'exported_at': datetime.now().isoformat(),
                    'count': len(data),
                    'students': data,
                },
                ensure_ascii=False,
                indent=2
            ),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M")}.json"'

        return response
