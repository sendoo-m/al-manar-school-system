from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from django.urls import reverse, path
from django.utils.html import format_html
from django.db import models, transaction
from django.forms import TextInput, Textarea
from django.contrib.admin import SimpleListFilter
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.exceptions import ValidationError
import csv
import io
import openpyxl
from datetime import datetime

from .models import Student, UserProfile, ArchiveStudent
from school_settings.models import AcademicYear as SettingsAcademicYear, GradeLevel


# ===================================
# تخصيص رأس لوحة الإدارة
# ===================================

admin.site.site_header = "نظام إدارة مدرسة المنار"
admin.site.site_title = "إدارة مدرسة المنار"
admin.site.index_title = _('لوحة التحكم الرئيسية')


# ===================================
# فلاتر لوحة الإدارة
# ===================================

class GenderFilter(SimpleListFilter):
    title = _('الجنس')
    parameter_name = 'gender'

    def lookups(self, request, model_admin):
        return (('M', _('ذكر')), ('F', _('أنثى')))

    def queryset(self, request, queryset):
        if self.value() == 'M':
            return queryset.filter(gender='M')
        if self.value() == 'F':
            return queryset.filter(gender='F')
        return queryset


class StudentTypeFilter(SimpleListFilter):
    title = _('نوع الطالب')
    parameter_name = 'student_type'

    def lookups(self, request, model_admin):
        return Student.STUDENT_TYPE_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(student_type=self.value())
        return queryset


class EnrollmentStatusFilter(SimpleListFilter):
    title = _('حالة القيد')
    parameter_name = 'enrollment_status'

    def lookups(self, request, model_admin):
        return Student.ENROLLMENT_STATUS_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(enrollment_status=self.value())
        return queryset


class ReligionFilter(SimpleListFilter):
    title = _('الديانة')
    parameter_name = 'religion'

    def lookups(self, request, model_admin):
        return (
            ('MUSLIM', _('مسلم')),
            ('CHRISTIAN', _('مسيحي')),
            ('OTHER', _('أخرى')),
        )

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(religion=self.value())
        return queryset


class IntegrationStudentFilter(SimpleListFilter):
    title = _('طلاب الدمج')
    parameter_name = 'integration_student'

    def lookups(self, request, model_admin):
        return (('yes', _('طلاب دمج')), ('no', _('غير دمج')))

    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.filter(is_integration_student=True)
        if self.value() == 'no':
            return queryset.filter(is_integration_student=False)
        return queryset


class StaffChildFilter(SimpleListFilter):
    title = _('أبناء العاملين')
    parameter_name = 'staff_child'

    def lookups(self, request, model_admin):
        return (('yes', _('من أبناء العاملين')), ('no', _('ليس من أبناء العاملين')))

    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.filter(is_staff_child=True)
        if self.value() == 'no':
            return queryset.filter(is_staff_child=False)
        return queryset


class FinancialStatusFilter(SimpleListFilter):
    title = _('الحالة المالية')
    parameter_name = 'financial_status'

    def lookups(self, request, model_admin):
        return (
            ('paid', _('مسدد بالكامل')),
            ('partial', _('مسدد جزئياً')),
            ('owing', _('مستحق عليه')),
        )

    def queryset(self, request, queryset):
        if self.value() == 'paid':
            return queryset.filter(total_owed__lte=0)
        if self.value() == 'partial':
            return queryset.filter(total_owed__gt=0, total_owed__lt=models.F('total_fees'))
        if self.value() == 'owing':
            return queryset.filter(total_owed__gte=models.F('total_fees'))
        return queryset


class AgeRangeFilter(SimpleListFilter):
    title = _('الفئة العمرية')
    parameter_name = 'age_range'

    def lookups(self, request, model_admin):
        return (
            ('3-6', _('3-6 سنوات (حضانة)')),
            ('6-12', _('6-12 سنة (ابتدائي)')),
            ('12-15', _('12-15 سنة (إعدادي)')),
            ('15-18', _('15-18 سنة (ثانوي)')),
        )

    def queryset(self, request, queryset):
        if self.value() == '3-6':
            return queryset.filter(age__gte=3, age__lte=6)
        if self.value() == '6-12':
            return queryset.filter(age__gte=6, age__lte=12)
        if self.value() == '12-15':
            return queryset.filter(age__gte=12, age__lte=15)
        if self.value() == '15-18':
            return queryset.filter(age__gte=15, age__lte=18)
        return queryset


# ===================================
# أدوات مساعدة للاستيراد
# ===================================

def get_text_value(data, *keys):
    for key in keys:
        value = data.get(key)
        if value is not None and str(value).strip() not in ['', 'None']:
            return str(value).strip()
    return ''


def get_bool_value(data, *keys):
    value = get_text_value(data, *keys).strip().lower()
    true_values = ['1', 'true', 'yes', 'y', 'نعم', 'صح', 'صحيح', 'موجود']
    false_values = ['0', 'false', 'no', 'n', 'لا', 'خطأ', 'غير موجود']

    if value in true_values:
        return True
    if value in false_values:
        return False
    return False


def get_choice_value(value, choices, default=''):
    value = str(value or '').strip()
    if not value:
        return default

    choices_dict = dict(choices)
    labels_to_values = {str(label).strip(): key for key, label in choices}

    if value in choices_dict:
        return value
    if value in labels_to_values:
        return labels_to_values[value]
    return default


def parse_date_value(value):
    if not value:
        return None

    if hasattr(value, 'date'):
        try:
            return value.date()
        except Exception:
            pass

    value = str(value).strip()
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def parse_decimal_value(value):
    try:
        if value in [None, '']:
            return 0
        return float(value)
    except (ValueError, TypeError):
        return 0


# ===================================
# إجراءات مخصصة
# ===================================

def export_students_to_csv(modeladmin, request, queryset):
    """تصدير الطلاب المحددين إلى ملف CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(modeladmin.get_export_headers())

    for student in queryset:
        writer.writerow(modeladmin.student_to_export_row(student))

    return response


export_students_to_csv.short_description = "تصدير الطلاب المحددين إلى CSV"


def archive_selected_students(modeladmin, request, queryset):
    """أرشفة الطلاب المحددين"""
    archived_count = 0

    for student in queryset:
        ArchiveStudent.objects.create(
            archive_name=student.name,
            archive_national_number=student.national_number or '',
            archive_passport_number=getattr(student, 'passport_number', '') or '',
            archive_student_type=student.get_student_type_display() if hasattr(student, 'get_student_type_display') else '',
            archive_nationality=getattr(student, 'nationality', '') or '',
            archive_religion=student.get_religion_display() if hasattr(student, 'get_religion_display') else '',
            archive_age=student.age or 0,
            archive_gender=student.gender,
            archive_date_of_birth=student.date_of_birth,

            archive_academic_year=str(student.academic_year) if student.academic_year else "غير محدد",
            archive_grade_level=student.grade_name if hasattr(student, 'grade_name') else "غير محدد",
            archive_education_level=student.education_level_name if hasattr(student, 'education_level_name') else "غير محدد",
            archive_enrollment_status=student.get_enrollment_status_display() if hasattr(student, 'get_enrollment_status_display') else '',
            archive_transferred_from_school=getattr(student, 'transferred_from_school', '') or '',
            archive_transferred_to_school=getattr(student, 'transferred_to_school', '') or '',

            archive_is_integration_student=getattr(student, 'is_integration_student', False),
            archive_disability_type=getattr(student, 'disability_type', '') or '',
            archive_subject_exemptions=student.get_subject_exemptions_display() if hasattr(student, 'get_subject_exemptions_display') else '',

            archive_total_payments=student.total_payments or 0,
            archive_total_fees=student.total_fees or 0,
            archive_total_owed=student.total_owed or 0,

            archive_parent_name=student.parent_name or '',
            archive_parent_phone=student.parent_phone or '',
            archive_father_job=getattr(student, 'father_job', '') or '',
            archive_educational_guardian=student.get_educational_guardian_display() if hasattr(student, 'get_educational_guardian_display') else '',
            archive_educational_guardian_name=getattr(student, 'educational_guardian_name', '') or '',
            archive_is_staff_child=getattr(student, 'is_staff_child', False),
            archive_staff_parent_name=getattr(student, 'staff_parent_name', '') or '',
            archive_staff_parent_job=getattr(student, 'staff_parent_job', '') or '',

            archived_reason='أرشفة من لوحة الإدارة',
            archived_by=request.user if request.user.is_authenticated else None,
        )
        archived_count += 1

    queryset.delete()
    modeladmin.message_user(request, f'تم أرشفة {archived_count} طالب وحذفهم من النظام الحالي.')


archive_selected_students.short_description = "أرشفة الطلاب المحددين"

# ===================================
# Mixin للاستيراد والتصدير
# ===================================

# class StudentImportExportMixin:
#     """Mixin لإضافة وظائف الاستيراد والتصدير للطلاب"""

#     def get_urls(self):
#         urls = super().get_urls()
#         custom_urls = [
#             path('export/', self.admin_site.admin_view(self.export_students), name='students_student_export'),
#             path('import/', self.admin_site.admin_view(self.import_students_view), name='students_student_import'),
#             path('import/process/', self.admin_site.admin_view(self.process_import), name='students_student_import_process'),
#             path('export/template/', self.admin_site.admin_view(self.download_template), name='students_student_export_template'),
#         ]
#         return custom_urls + urls

#     def get_export_headers(self):
#         return [
#             'الاسم*',
#             'نوع الطالب',
#             'الرقم القومي',
#             'رقم جواز السفر',
#             'الجنسية',
#             'الديانة',
#             'العمر',
#             'النوع (M/F)',
#             'تاريخ الميلاد (YYYY-MM-DD)',
#             'رقم الهاتف',
#             'العنوان',
#             'العام الدراسي',
#             'المرحلة التعليمية',
#             'الصف الدراسي',
#             'حالة القيد',
#             'محول من مدرسة',
#             'محول إلى مدرسة',
#             'طالب دمج',
#             'نوع الإعاقة',
#             'إعفاء من العربي',
#             'إعفاء من الإنجليزي',
#             'إعفاء من الفرنسي',
#             'إعفاءات أخرى',
#             'من أبناء العاملين',
#             'اسم الموظف',
#             'وظيفة الموظف',
#             'اسم ولي الأمر',
#             'هاتف ولي الأمر',
#             'بريد ولي الأمر',
#             'وظيفة الأب',
#             'صاحب الولاية التعليمية',
#             'اسم صاحب الولاية التعليمية',
#             'هاتف صاحب الولاية التعليمية',
#             'إجمالي المصروفات',
#             'إجمالي المدفوعات',
#             'المستحقات',
#             'تاريخ الإنشاء',
#             'الحالة',
#         ]

#     def get_import_headers(self):
#         return [
#             'الاسم*',
#             'نوع الطالب',
#             'الرقم القومي',
#             'رقم جواز السفر',
#             'الجنسية',
#             'الديانة',
#             'العمر',
#             'النوع (M/F)',
#             'تاريخ الميلاد (YYYY-MM-DD)',
#             'رقم الهاتف',
#             'العنوان',
#             'العام الدراسي',
#             'المرحلة التعليمية',
#             'الصف الدراسي',
#             'حالة القيد',
#             'محول من مدرسة',
#             'محول إلى مدرسة',
#             'طالب دمج',
#             'نوع الإعاقة',
#             'إعفاء من العربي',
#             'إعفاء من الإنجليزي',
#             'إعفاء من الفرنسي',
#             'إعفاءات أخرى',
#             'من أبناء العاملين',
#             'اسم الموظف',
#             'وظيفة الموظف',
#             'اسم ولي الأمر',
#             'هاتف ولي الأمر',
#             'بريد ولي الأمر',
#             'وظيفة الأب',
#             'صاحب الولاية التعليمية',
#             'اسم صاحب الولاية التعليمية',
#             'هاتف صاحب الولاية التعليمية',
#             'إجمالي المصروفات',
#             'إجمالي المدفوعات',
#         ]

#     def student_to_export_row(self, student):
#         return [
#             student.name,
#             student.get_student_type_display() if hasattr(student, 'get_student_type_display') else '',
#             student.national_number or '',
#             getattr(student, 'passport_number', '') or '',
#             getattr(student, 'nationality', '') or '',
#             student.get_religion_display() if hasattr(student, 'get_religion_display') else '',
#             student.age or '',
#             student.gender if student.gender else '',
#             student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
#             student.phone_number or '',
#             student.address or '',
#             student.academic_year.name if student.academic_year else '',
#             student.grade_level.education_level.name if (student.grade_level and getattr(student.grade_level, 'education_level', None)) else '',
#             student.grade_level.name if student.grade_level else '',
#             student.get_enrollment_status_display() if hasattr(student, 'get_enrollment_status_display') else '',
#             getattr(student, 'transferred_from_school', '') or '',
#             getattr(student, 'transferred_to_school', '') or '',
#             'نعم' if getattr(student, 'is_integration_student', False) else 'لا',
#             getattr(student, 'disability_type', '') or '',
#             'نعم' if getattr(student, 'exempt_from_arabic', False) else 'لا',
#             'نعم' if getattr(student, 'exempt_from_english', False) else 'لا',
#             'نعم' if getattr(student, 'exempt_from_french', False) else 'لا',
#             getattr(student, 'other_subject_exemptions', '') or '',
#             'نعم' if getattr(student, 'is_staff_child', False) else 'لا',
#             getattr(student, 'staff_parent_name', '') or '',
#             getattr(student, 'staff_parent_job', '') or '',
#             student.parent_name or '',
#             student.parent_phone or '',
#             student.parent_email or '',
#             getattr(student, 'father_job', '') or '',
#             student.get_educational_guardian_display() if hasattr(student, 'get_educational_guardian_display') else '',
#             getattr(student, 'educational_guardian_name', '') or '',
#             getattr(student, 'educational_guardian_phone', '') or '',
#             float(student.total_fees or 0),
#             float(student.total_payments or 0),
#             float(student.total_owed or 0),
#             student.created_at.strftime('%Y-%m-%d %H:%M:%S') if student.created_at else '',
#             'نشط' if student.is_active else 'غير نشط',
#         ]

#     def export_students(self, request):
#         export_format = request.GET.get('format', 'excel')
#         grade_filter = request.GET.get('grade')

#         queryset = Student.objects.filter(is_active=True).select_related(
#             'academic_year',
#             'grade_level__education_level'
#         )

#         if grade_filter:
#             queryset = queryset.filter(grade_level_id=grade_filter)

#         if export_format == 'csv':
#             return self._export_csv(queryset)
#         if export_format == 'excel':
#             return self._export_excel(queryset)
#         if export_format == 'json':
#             return self._export_json(queryset)

#         messages.error(request, 'صيغة التصدير غير مدعومة')
#         return redirect('admin:students_student_changelist')

#     def _export_csv(self, queryset):
#         response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
#         response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
#         response.write('\ufeff')

#         writer = csv.writer(response)
#         writer.writerow(self.get_export_headers())

#         for student in queryset:
#             writer.writerow(self.student_to_export_row(student))

#         return response

#     def _export_excel(self, queryset):
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "بيانات الطلاب"
#         ws.sheet_view.rightToLeft = True

#         headers = self.get_export_headers()

#         for col, header in enumerate(headers, 1):
#             cell = ws.cell(row=1, column=col, value=header)
#             cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
#             cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#             cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

#         for row_num, student in enumerate(queryset, 2):
#             for col, value in enumerate(self.student_to_export_row(student), 1):
#                 cell = ws.cell(row=row_num, column=col, value=value)
#                 if col in [34, 35, 36]:
#                     cell.number_format = '#,##0.00'

#         ws.freeze_panes = 'A2'
#         ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(headers)).coordinate}'

#         for column in ws.columns:
#             max_length = 0
#             column_letter = column[0].column_letter
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except Exception:
#                     pass
#             ws.column_dimensions[column_letter].width = min(max_length + 2, 45)

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         response = HttpResponse(
#             output.read(),
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
#         return response

#     def _export_json(self, queryset):
#         data = []
#         for student in queryset:
#             data.append({
#                 'name': student.name,
#                 'student_type': getattr(student, 'student_type', ''),
#                 'national_number': student.national_number,
#                 'passport_number': getattr(student, 'passport_number', ''),
#                 'nationality': getattr(student, 'nationality', ''),
#                 'religion': getattr(student, 'religion', ''),
#                 'age': student.age,
#                 'gender': student.gender,
#                 'date_of_birth': student.date_of_birth.isoformat() if student.date_of_birth else None,
#                 'phone_number': student.phone_number,
#                 'address': student.address,
#                 'academic_year': student.academic_year.name if student.academic_year else None,
#                 'education_level': student.grade_level.education_level.name if (student.grade_level and getattr(student.grade_level, 'education_level', None)) else None,
#                 'grade_level': student.grade_level.name if student.grade_level else None,
#                 'enrollment_status': getattr(student, 'enrollment_status', ''),
#                 'transferred_from_school': getattr(student, 'transferred_from_school', ''),
#                 'transferred_to_school': getattr(student, 'transferred_to_school', ''),
#                 'is_integration_student': getattr(student, 'is_integration_student', False),
#                 'disability_type': getattr(student, 'disability_type', ''),
#                 'subject_exemptions': student.get_subject_exemptions_display() if hasattr(student, 'get_subject_exemptions_display') else '',
#                 'is_staff_child': getattr(student, 'is_staff_child', False),
#                 'parent_name': student.parent_name,
#                 'parent_phone': student.parent_phone,
#                 'parent_email': student.parent_email,
#                 'father_job': getattr(student, 'father_job', ''),
#                 'educational_guardian': getattr(student, 'educational_guardian', ''),
#                 'educational_guardian_name': getattr(student, 'educational_guardian_name', ''),
#                 'educational_guardian_phone': getattr(student, 'educational_guardian_phone', ''),
#                 'total_fees': float(student.total_fees or 0),
#                 'total_payments': float(student.total_payments or 0),
#                 'total_owed': float(student.total_owed or 0),
#                 'created_at': student.created_at.isoformat() if student.created_at else None,
#                 'is_active': student.is_active,
#             })

#         response = JsonResponse({
#             'students': data,
#             'export_date': datetime.now().isoformat(),
#             'total_count': len(data)
#         }, json_dumps_params={'ensure_ascii': False, 'indent': 2})
#         response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
#         return response

#     def download_template(self, request):
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "قالب استيراد الطلاب"
#         ws.sheet_view.rightToLeft = True

#         headers = self.get_import_headers()

#         for col, header in enumerate(headers, 1):
#             cell = ws.cell(row=1, column=col, value=header)
#             cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
#             cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#             cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

#         sample_rows = [
#             [
#                 'أحمد محمد علي',
#                 'طالب عادي',
#                 '30012010012345',
#                 '',
#                 'مصري',
#                 'مسلم',
#                 '14',
#                 'M',
#                 '2010-01-01',
#                 '01234567890',
#                 'القاهرة',
#                 '2024-2025',
#                 'الإعدادية',
#                 'الثاني الإعدادي',
#                 'مستجد',
#                 '',
#                 '',
#                 'لا',
#                 '',
#                 'لا',
#                 'لا',
#                 'لا',
#                 '',
#                 'لا',
#                 '',
#                 '',
#                 'محمد علي أحمد',
#                 '01098765432',
#                 'parent@email.com',
#                 'محاسب',
#                 'الأب',
#                 '',
#                 '',
#                 '15000',
#                 '5000',
#             ],
#             [
#                 'يوسف جون سميث',
#                 'وافد',
#                 '',
#                 'A12345678',
#                 'أمريكي',
#                 'مسيحي',
#                 '8',
#                 'M',
#                 '2016-05-10',
#                 '01000000002',
#                 'الإسماعيلية',
#                 '2024-2025',
#                 'الابتدائية',
#                 'الثاني الابتدائي',
#                 'محول',
#                 'مدرسة دولية سابقة',
#                 '',
#                 'لا',
#                 '',
#                 'لا',
#                 'لا',
#                 'لا',
#                 '',
#                 'لا',
#                 '',
#                 '',
#                 'John Smith',
#                 '01111111113',
#                 'john@example.com',
#                 'مهندس',
#                 'الأب',
#                 '',
#                 '',
#                 '18000',
#                 '10000',
#             ],
#         ]

#         for row_num, row_data in enumerate(sample_rows, 2):
#             for col, value in enumerate(row_data, 1):
#                 ws.cell(row=row_num, column=col, value=value)

#         instructions = [
#             "تعليمات الاستيراد:",
#             "1. الاسم هو الحقل الإجباري الوحيد.",
#             "2. الرقم القومي اختياري، وإذا تم إدخاله يجب ألا يكون مكرراً.",
#             "3. رقم جواز السفر اختياري ويستخدم غالباً للوافدين، ويجب ألا يكون مكرراً إذا تم إدخاله.",
#             "4. نوع الطالب يقبل: طالب عادي أو وافد.",
#             "5. النوع يقبل: M أو F.",
#             "6. القيم المنطقية تقبل: نعم / لا / 1 / 0.",
#             "7. المرحلة التعليمية يمكن كتابتها مثل: الابتدائية / الإعدادية / الثانوية.",
#             "8. الصف الدراسي يمكن كتابته بالاسم مثل: الثاني الإعدادي.",
#             "9. عند تكرار اسم الصف في أكثر من مرحلة يجب كتابة المرحلة التعليمية.",
#             "10. العام الدراسي اختياري، وإذا تُرك فارغاً يمكن للنظام تركه فارغاً أو استخدام الإعداد الافتراضي حسب البيانات المتاحة.",
#             "11. تاريخ الميلاد يفضل بصيغة YYYY-MM-DD.",
#             "12. لا تغير أسماء الأعمدة في الصف الأول.",
#             "13. احذف سطور التعليمات قبل الاستيراد.",
#         ]

#         start_row = len(sample_rows) + 4
#         for index, instruction in enumerate(instructions):
#             cell = ws.cell(row=start_row + index, column=1, value=instruction)
#             if index == 0:
#                 cell.font = openpyxl.styles.Font(bold=True, color="D32F2F")
#             else:
#                 cell.font = openpyxl.styles.Font(color="666666")

#         ws.freeze_panes = 'A2'
#         ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(headers)).coordinate}'

#         for column in ws.columns:
#             max_length = 0
#             column_letter = column[0].column_letter
#             for cell in column:
#                 try:
#                     max_length = max(max_length, len(str(cell.value)))
#                 except Exception:
#                     pass
#             ws.column_dimensions[column_letter].width = min(max_length + 2, 35)

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         response = HttpResponse(
#             output.read(),
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
#         return response

#     def import_students_view(self, request):
#         grade_levels = GradeLevel.objects.filter(
#             is_active=True
#         ).select_related(
#             'education_level'
#         ).order_by(
#             'education_level__order',
#             'order'
#         )

#         academic_years = SettingsAcademicYear.objects.filter(is_active=True).order_by('-start_date')

#         context = {
#             'title': 'استيراد بيانات الطلاب',
#             'grade_levels': grade_levels,
#             'academic_years': academic_years,
#             'opts': Student._meta,
#         }
#         return render(request, 'admin/students/student/import_students.html', context)

#     def process_import(self, request):
#         if request.method != 'POST':
#             return redirect('admin:students_student_import')

#         file = request.FILES.get('import_file')
#         if not file:
#             messages.error(request, 'يرجى اختيار ملف للاستيراد')
#             return redirect('admin:students_student_import')

#         try:
#             if file.name.endswith('.xlsx'):
#                 success_count, errors = self._process_excel_import(file, request)
#             elif file.name.endswith('.csv'):
#                 success_count, errors = self._process_csv_import(file, request)
#             else:
#                 messages.error(request, 'صيغة الملف غير مدعومة. يرجى استخدام Excel أو CSV')
#                 return redirect('admin:students_student_import')

#             if success_count > 0:
#                 messages.success(request, f'تم استيراد {success_count} طالب بنجاح')

#             if errors:
#                 error_msg = f'حدثت {len(errors)} أخطاء أثناء الاستيراد:\n' + '\n'.join(errors[:8])
#                 if len(errors) > 8:
#                     error_msg += f'\n... و {len(errors) - 8} أخطاء أخرى'
#                 messages.warning(request, error_msg)

#         except Exception as e:
#             messages.error(request, f'حدث خطأ أثناء معالجة الملف: {str(e)}')

#         return redirect('admin:students_student_changelist')

#     def _is_instruction_row(self, first_cell_value):
#         value = str(first_cell_value or '').strip()
#         if not value:
#             return False

#         if value.startswith('تعليمات الاستيراد'):
#             return True

#         instruction_prefixes = [f'{i}.' for i in range(1, 20)]
#         return any(value.startswith(prefix) for prefix in instruction_prefixes)

#     def _process_excel_import(self, file, request):
#         wb = openpyxl.load_workbook(file)
#         ws = wb.active
#         success_count = 0
#         errors = []
#         headers = [cell.value for cell in ws[1]]

#         with transaction.atomic():
#             for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
#                 if not any(row):
#                     continue

#                 first_cell = str(row[0]).strip() if row and row[0] is not None else ''
#                 if self._is_instruction_row(first_cell):
#                     continue

#                 try:
#                     student_data = dict(zip(headers, row))
#                     student = self._create_student_from_data(student_data, row_num)
#                     if student:
#                         success_count += 1
#                 except ValidationError as e:
#                     errors.append(f'الصف {row_num}: {", ".join(e.messages)}')
#                 except Exception as e:
#                     errors.append(f'الصف {row_num}: {str(e)}')

#         return success_count, errors

#     def _process_csv_import(self, file, request):
#         success_count = 0
#         errors = []
#         file_content = file.read().decode('utf-8-sig')
#         csv_reader = csv.DictReader(io.StringIO(file_content))

#         with transaction.atomic():
#             for row_num, row in enumerate(csv_reader, 2):
#                 if not any(str(v).strip() for v in row.values() if v is not None):
#                     continue

#                 first_value = next(iter(row.values()), '')
#                 if self._is_instruction_row(first_value):
#                     continue

#                 try:
#                     student = self._create_student_from_data(row, row_num)
#                     if student:
#                         success_count += 1
#                 except ValidationError as e:
#                     errors.append(f'الصف {row_num}: {", ".join(e.messages)}')
#                 except Exception as e:
#                     errors.append(f'الصف {row_num}: {str(e)}')

#         return success_count, errors

#     def _find_grade_level(self, grade_name='', education_level_name=''):
#         grade_name = str(grade_name or '').strip()
#         education_level_name = str(education_level_name or '').strip()

#         if not grade_name:
#             return None

#         queryset = GradeLevel.objects.filter(is_active=True).select_related('education_level')

#         if education_level_name:
#             exact_matches = queryset.filter(
#                 name__iexact=grade_name,
#                 education_level__name__iexact=education_level_name
#             )
#             if exact_matches.count() == 1:
#                 return exact_matches.first()
#             if exact_matches.count() > 1:
#                 raise ValidationError(
#                     f"يوجد أكثر من صف مطابق للاسم '{grade_name}' داخل المرحلة '{education_level_name}'"
#                 )

#             contains_matches = queryset.filter(
#                 name__icontains=grade_name,
#                 education_level__name__iexact=education_level_name
#             )
#             if contains_matches.count() == 1:
#                 return contains_matches.first()
#             if contains_matches.count() > 1:
#                 raise ValidationError(
#                     f"يوجد أكثر من صف مشابه للاسم '{grade_name}' داخل المرحلة '{education_level_name}'"
#                 )

#             raise ValidationError(
#                 f"تعذر العثور على الصف الدراسي '{grade_name}' داخل المرحلة التعليمية '{education_level_name}'"
#             )

#         exact_matches = queryset.filter(name__iexact=grade_name)
#         if exact_matches.count() == 1:
#             return exact_matches.first()
#         if exact_matches.count() > 1:
#             raise ValidationError(
#                 f"اسم الصف الدراسي '{grade_name}' مكرر في أكثر من مرحلة، يرجى تحديد المرحلة التعليمية"
#             )

#         contains_matches = queryset.filter(name__icontains=grade_name)
#         if contains_matches.count() == 1:
#             return contains_matches.first()
#         if contains_matches.count() > 1:
#             raise ValidationError(
#                 f"اسم الصف الدراسي '{grade_name}' غير محدد بدقة، يرجى كتابة المرحلة التعليمية أيضاً"
#             )

#         raise ValidationError(f"تعذر العثور على الصف الدراسي '{grade_name}'")

#     def _find_academic_year(self, year_name=''):
#         year_name = str(year_name or '').strip()
#         if not year_name:
#             return None

#         exact_match = SettingsAcademicYear.objects.filter(
#             name__iexact=year_name,
#             is_active=True
#         ).first()
#         if exact_match:
#             return exact_match

#         return SettingsAcademicYear.objects.filter(
#             name__icontains=year_name,
#             is_active=True
#         ).first()

#     def _normalize_gender(self, value):
#         value = str(value or '').strip().upper()
#         if value in ['M', 'MALE', 'ذكر']:
#             return 'M'
#         if value in ['F', 'FEMALE', 'أنثى', 'انثى']:
#             return 'F'
#         return ''

#     def download_reference_data(self, request):
#         wb = openpyxl.Workbook()

#         # =========================================
#         # الشيت 1: قالب الاستيراد
#         # =========================================
#         ws_template = wb.active
#         ws_template.title = "قالب الاستيراد"
#         ws_template.sheet_view.rightToLeft = True

#         headers = self.get_import_headers()

#         header_fill = openpyxl.styles.PatternFill(
#             start_color="4472C4",
#             end_color="4472C4",
#             fill_type="solid"
#         )
#         header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
#         center_alignment = openpyxl.styles.Alignment(
#             horizontal="center",
#             vertical="center",
#             wrap_text=True
#         )

#         for col, header in enumerate(headers, 1):
#             cell = ws_template.cell(row=1, column=col, value=header)
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = center_alignment

#         sample_rows = [
#             [
#                 'أحمد محمد علي',
#                 'طالب عادي',
#                 '30012010012345',
#                 '',
#                 'مصري',
#                 'مسلم',
#                 '14',
#                 'M',
#                 '2010-01-01',
#                 '01234567890',
#                 'القاهرة',
#                 '2024-2025',
#                 'الإعدادية',
#                 'الثاني الإعدادي',
#                 'مستجد',
#                 '',
#                 '',
#                 'لا',
#                 '',
#                 'لا',
#                 'لا',
#                 'لا',
#                 '',
#                 'لا',
#                 '',
#                 '',
#                 'محمد علي أحمد',
#                 '01098765432',
#                 'parent@email.com',
#                 'محاسب',
#                 'الأب',
#                 '',
#                 '',
#                 '15000',
#                 '5000',
#             ]
#         ]

#         for row_num, row_data in enumerate(sample_rows, 2):
#             for col, value in enumerate(row_data, 1):
#                 cell = ws_template.cell(row=row_num, column=col, value=value)
#                 cell.alignment = center_alignment

#         instructions = [
#             "تعليمات سريعة:",
#             "1. استخدم نفس أسماء الأعمدة بدون تغيير.",
#             "2. انسخ العام الدراسي والمرحلة التعليمية والصف الدراسي من الشيتات المرجعية.",
#             "3. إذا كان اسم الصف مكرراً في أكثر من مرحلة، يجب كتابة المرحلة التعليمية.",
#             "4. الاسم هو الحقل الإجباري الوحيد.",
#             "5. الرقم القومي أو جواز السفر يجب ألا يكون مكرراً إذا تم إدخاله.",
#             "6. النوع يقبل: M / F أو ذكر / أنثى.",
#             "7. تاريخ الميلاد يفضل بصيغة YYYY-MM-DD.",
#         ]

#         start_row = len(sample_rows) + 4
#         for index, instruction in enumerate(instructions):
#             cell = ws_template.cell(row=start_row + index, column=1, value=instruction)
#             if index == 0:
#                 cell.font = openpyxl.styles.Font(bold=True, color="D32F2F")
#             else:
#                 cell.font = openpyxl.styles.Font(color="666666")

#         ws_template.freeze_panes = 'A2'
#         ws_template.auto_filter.ref = f'A1:{ws_template.cell(row=1, column=len(headers)).coordinate}'

#         # =========================================
#         # الشيت 2: الأعوام الدراسية
#         # =========================================
#         ws_years = wb.create_sheet(title="الاعوام الدراسية")
#         ws_years.sheet_view.rightToLeft = True

#         year_headers = ['م', 'العام الدراسي', 'نشط؟']
#         for col, header in enumerate(year_headers, 1):
#             cell = ws_years.cell(row=1, column=col, value=header)
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = center_alignment

#         academic_years = SettingsAcademicYear.objects.all().order_by('-start_date', '-id')
#         for row_num, academic_year in enumerate(academic_years, 2):
#             ws_years.cell(row=row_num, column=1, value=row_num - 1)
#             ws_years.cell(row=row_num, column=2, value=getattr(academic_year, 'name', str(academic_year)))
#             ws_years.cell(row=row_num, column=3, value='نعم' if getattr(academic_year, 'is_active', False) else 'لا')

#         ws_years.freeze_panes = 'A2'
#         ws_years.auto_filter.ref = f'A1:C{max(ws_years.max_row, 2)}'

#         # =========================================
#         # الشيت 3: المراحل والصفوف
#         # =========================================
#         ws_grades = wb.create_sheet(title="المراحل والصفوف")
#         ws_grades.sheet_view.rightToLeft = True

#         grade_headers = ['م', 'المرحلة التعليمية', 'الصف الدراسي', 'نشط؟']
#         for col, header in enumerate(grade_headers, 1):
#             cell = ws_grades.cell(row=1, column=col, value=header)
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = center_alignment

#         grade_levels = GradeLevel.objects.select_related('education_level').all().order_by(
#             'education_level__order',
#             'order',
#             'name'
#         )

#         for row_num, grade_level in enumerate(grade_levels, 2):
#             ws_grades.cell(row=row_num, column=1, value=row_num - 1)
#             ws_grades.cell(
#                 row=row_num,
#                 column=2,
#                 value=grade_level.education_level.name if getattr(grade_level, 'education_level', None) else ''
#             )
#             ws_grades.cell(row=row_num, column=3, value=grade_level.name)
#             ws_grades.cell(row=row_num, column=4, value='نعم' if getattr(grade_level, 'is_active', False) else 'لا')

#         ws_grades.freeze_panes = 'A2'
#         ws_grades.auto_filter.ref = f'A1:D{max(ws_grades.max_row, 2)}'

#         # =========================================
#         # الشيت 4: القيم المسموح بها
#         # =========================================
#         ws_choices = wb.create_sheet(title="قيم مسموح بها")
#         ws_choices.sheet_view.rightToLeft = True

#         choice_headers = ['نوع البيان', 'القيمة المعروضة', 'القيمة الداخلية']
#         for col, header in enumerate(choice_headers, 1):
#             cell = ws_choices.cell(row=1, column=col, value=header)
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = center_alignment

#         row_num = 2

#         for value, label in getattr(Student, 'STUDENT_TYPE_CHOICES', []):
#             ws_choices.cell(row=row_num, column=1, value='نوع الطالب')
#             ws_choices.cell(row=row_num, column=2, value=str(label))
#             ws_choices.cell(row=row_num, column=3, value=str(value))
#             row_num += 1

#         for value, label in getattr(Student, 'RELIGION_CHOICES', []):
#             ws_choices.cell(row=row_num, column=1, value='الديانة')
#             ws_choices.cell(row=row_num, column=2, value=str(label))
#             ws_choices.cell(row=row_num, column=3, value=str(value))
#             row_num += 1

#         for value, label in getattr(Student, 'ENROLLMENT_STATUS_CHOICES', []):
#             ws_choices.cell(row=row_num, column=1, value='حالة القيد')
#             ws_choices.cell(row=row_num, column=2, value=str(label))
#             ws_choices.cell(row=row_num, column=3, value=str(value))
#             row_num += 1

#         for value, label in getattr(Student, 'EDUCATIONAL_GUARDIAN_CHOICES', []):
#             ws_choices.cell(row=row_num, column=1, value='صاحب الولاية التعليمية')
#             ws_choices.cell(row=row_num, column=2, value=str(label))
#             ws_choices.cell(row=row_num, column=3, value=str(value))
#             row_num += 1

#         extra_values = [
#             ('النوع', 'M', 'M'),
#             ('النوع', 'F', 'F'),
#             ('النوع', 'ذكر', 'M'),
#             ('النوع', 'أنثى', 'F'),
#             ('القيم المنطقية', 'نعم', 'True'),
#             ('القيم المنطقية', 'لا', 'False'),
#             ('القيم المنطقية', '1', 'True'),
#             ('القيم المنطقية', '0', 'False'),
#         ]

#         for item_type, display_value, internal_value in extra_values:
#             ws_choices.cell(row=row_num, column=1, value=item_type)
#             ws_choices.cell(row=row_num, column=2, value=display_value)
#             ws_choices.cell(row=row_num, column=3, value=internal_value)
#             row_num += 1

#         ws_choices.freeze_panes = 'A2'
#         ws_choices.auto_filter.ref = f'A1:C{max(ws_choices.max_row, 2)}'

#         # =========================================
#         # ضبط عرض الأعمدة لكل الشيتات
#         # =========================================
#         for ws in wb.worksheets:
#             for column in ws.columns:
#                 max_length = 0
#                 column_letter = column[0].column_letter
#                 for cell in column:
#                     try:
#                         value_length = len(str(cell.value)) if cell.value else 0
#                         if value_length > max_length:
#                             max_length = value_length
#                     except Exception:
#                         pass
#                 ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         response = HttpResponse(
#             output.read(),
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="student_import_reference_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
#         return response


#     def _create_student_from_data(self, data, row_num):
#         name = get_text_value(data, 'الاسم*', 'الاسم', 'name')
#         if not name:
#             raise ValidationError('اسم الطالب مطلوب')

#         national_number = get_text_value(data, 'الرقم القومي*', 'الرقم القومي', 'national_number')
#         passport_number = get_text_value(data, 'رقم جواز السفر', 'passport_number')

#         if national_number and Student.objects.filter(national_number=national_number).exists():
#             raise ValidationError(f'الرقم القومي {national_number} موجود مسبقاً')

#         if passport_number and Student.objects.filter(passport_number=passport_number).exists():
#             raise ValidationError(f'رقم جواز السفر {passport_number} موجود مسبقاً')

#         age = get_text_value(data, 'العمر', 'age')
#         try:
#             age = int(age) if age else None
#         except (ValueError, TypeError):
#             age = None

#         gender = self._normalize_gender(
#             get_text_value(data, 'النوع (M/F)', 'النوع', 'gender')
#         )

#         date_of_birth = parse_date_value(
#             get_text_value(data, 'تاريخ الميلاد (YYYY-MM-DD)', 'تاريخ الميلاد', 'date_of_birth')
#         )

#         education_level_name = get_text_value(data, 'المرحلة التعليمية', 'education_level')
#         grade_name = get_text_value(data, 'الصف الدراسي', 'grade_level')
#         academic_year_name = get_text_value(data, 'العام الدراسي', 'academic_year')

#         grade_level = self._find_grade_level(
#             grade_name=grade_name,
#             education_level_name=education_level_name
#         ) if grade_name else None

#         academic_year = self._find_academic_year(academic_year_name)

#         student_type = get_choice_value(
#             get_text_value(data, 'نوع الطالب', 'student_type'),
#             Student.STUDENT_TYPE_CHOICES,
#             default='REGULAR'
#         )

#         religion = get_choice_value(
#             get_text_value(data, 'الديانة', 'religion'),
#             Student.RELIGION_CHOICES,
#             default=''
#         )

#         enrollment_status = get_choice_value(
#             get_text_value(data, 'حالة القيد', 'enrollment_status'),
#             Student.ENROLLMENT_STATUS_CHOICES,
#             default='NEW'
#         )

#         educational_guardian = get_choice_value(
#             get_text_value(data, 'صاحب الولاية التعليمية', 'educational_guardian'),
#             Student.EDUCATIONAL_GUARDIAN_CHOICES,
#             default='FATHER'
#         )

#         total_fees = parse_decimal_value(get_text_value(data, 'إجمالي المصروفات', 'total_fees'))
#         total_payments = parse_decimal_value(get_text_value(data, 'إجمالي المدفوعات', 'total_payments'))
#         total_owed = total_fees - total_payments

#         student = Student.objects.create(
#             name=name,
#             student_type=student_type,
#             national_number=national_number or None,
#             passport_number=passport_number or None,
#             nationality=get_text_value(data, 'الجنسية', 'nationality'),
#             religion=religion,
#             age=age,
#             gender=gender,
#             date_of_birth=date_of_birth,
#             phone_number=get_text_value(data, 'رقم الهاتف', 'phone_number'),
#             address=get_text_value(data, 'العنوان', 'address'),
#             grade_level=grade_level,
#             academic_year=academic_year,
#             enrollment_status=enrollment_status,
#             transferred_from_school=get_text_value(data, 'محول من مدرسة', 'transferred_from_school'),
#             transferred_to_school=get_text_value(data, 'محول إلى مدرسة', 'transferred_to_school'),
#             is_integration_student=get_bool_value(data, 'طالب دمج', 'is_integration_student'),
#             disability_type=get_text_value(data, 'نوع الإعاقة', 'disability_type'),
#             exempt_from_arabic=get_bool_value(data, 'إعفاء من العربي', 'exempt_from_arabic'),
#             exempt_from_english=get_bool_value(data, 'إعفاء من الإنجليزي', 'exempt_from_english'),
#             exempt_from_french=get_bool_value(data, 'إعفاء من الفرنسي', 'exempt_from_french'),
#             other_subject_exemptions=get_text_value(data, 'إعفاءات أخرى', 'other_subject_exemptions'),
#             is_staff_child=get_bool_value(data, 'من أبناء العاملين', 'is_staff_child'),
#             staff_parent_name=get_text_value(data, 'اسم الموظف', 'staff_parent_name'),
#             staff_parent_job=get_text_value(data, 'وظيفة الموظف', 'staff_parent_job'),
#             parent_name=get_text_value(data, 'اسم ولي الأمر', 'parent_name'),
#             parent_phone=get_text_value(data, 'هاتف ولي الأمر', 'parent_phone'),
#             parent_email=get_text_value(data, 'بريد ولي الأمر', 'parent_email'),
#             father_job=get_text_value(data, 'وظيفة الأب', 'father_job'),
#             educational_guardian=educational_guardian,
#             educational_guardian_name=get_text_value(data, 'اسم صاحب الولاية التعليمية', 'educational_guardian_name'),
#             educational_guardian_phone=get_text_value(data, 'هاتف صاحب الولاية التعليمية', 'educational_guardian_phone'),
#             total_fees=total_fees,
#             total_payments=total_payments,
#             total_owed=total_owed,
#             is_active=True,
#         )

#         return student


class StudentImportExportMixin:
    """Mixin لإضافة وظائف الاستيراد والتصدير للطلاب"""

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export/', self.admin_site.admin_view(self.export_students), name='students_student_export'),
            path('import/', self.admin_site.admin_view(self.import_students_view), name='students_student_import'),
            path('import/process/', self.admin_site.admin_view(self.process_import), name='students_student_import_process'),
            path('export/template/', self.admin_site.admin_view(self.download_template), name='students_student_export_template'),
            path('export/reference/', self.admin_site.admin_view(self.download_reference_data), name='students_student_export_reference'),
        ]
        return custom_urls + urls

    def get_import_headers(self):
        return [
            'الاسم*',
            'نوع الطالب',
            'الرقم القومي',
            'رقم جواز السفر',
            'الجنسية',
            'الديانة',
            'العمر',
            'النوع',
            'تاريخ الميلاد',
            'رقم الهاتف',
            'العنوان',
            'العام الدراسي',
            'المرحلة التعليمية',
            'الصف الدراسي',
            'حالة القيد',
            'محول من مدرسة',
            'محول إلى مدرسة',
            'طالب دمج',
            'نوع الإعاقة',
            'إعفاء من العربي',
            'إعفاء من الإنجليزي',
            'إعفاء من الفرنسي',
            'إعفاءات أخرى',
            'من أبناء العاملين',
            'اسم الموظف',
            'وظيفة الموظف',
            'اسم ولي الأمر',
            'هاتف ولي الأمر',
            'بريد ولي الأمر',
            'وظيفة الأب',
            'صاحب الولاية التعليمية',
            'اسم صاحب الولاية التعليمية',
            'هاتف صاحب الولاية التعليمية',
            'إجمالي المصروفات',
            'إجمالي المدفوعات',
        ]

    def get_export_headers(self):
        return [
            'الاسم*',
            'نوع الطالب',
            'الرقم القومي',
            'رقم جواز السفر',
            'الجنسية',
            'الديانة',
            'العمر',
            'النوع',
            'تاريخ الميلاد',
            'رقم الهاتف',
            'العنوان',
            'العام الدراسي',
            'المرحلة التعليمية',
            'الصف الدراسي',
            'حالة القيد',
            'محول من مدرسة',
            'محول إلى مدرسة',
            'طالب دمج',
            'نوع الإعاقة',
            'إعفاء من العربي',
            'إعفاء من الإنجليزي',
            'إعفاء من الفرنسي',
            'إعفاءات أخرى',
            'من أبناء العاملين',
            'اسم الموظف',
            'وظيفة الموظف',
            'اسم ولي الأمر',
            'هاتف ولي الأمر',
            'بريد ولي الأمر',
            'وظيفة الأب',
            'صاحب الولاية التعليمية',
            'اسم صاحب الولاية التعليمية',
            'هاتف صاحب الولاية التعليمية',
            'إجمالي المصروفات',
            'إجمالي المدفوعات',
            'المستحقات',
            'تاريخ الإنشاء',
            'الحالة',
        ]

    def student_to_export_row(self, student):
        return [
            student.name,
            student.get_student_type_display() if hasattr(student, 'get_student_type_display') else '',
            student.national_number or '',
            getattr(student, 'passport_number', '') or '',
            getattr(student, 'nationality', '') or '',
            student.get_religion_display() if hasattr(student, 'get_religion_display') else '',
            student.age or '',
            student.gender if student.gender else '',
            student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
            student.phone_number or '',
            student.address or '',
            student.academic_year.name if student.academic_year else '',
            student.grade_level.education_level.name if (student.grade_level and getattr(student.grade_level, 'education_level', None)) else '',
            student.grade_level.name if student.grade_level else '',
            student.get_enrollment_status_display() if hasattr(student, 'get_enrollment_status_display') else '',
            getattr(student, 'transferred_from_school', '') or '',
            getattr(student, 'transferred_to_school', '') or '',
            'نعم' if getattr(student, 'is_integration_student', False) else 'لا',
            getattr(student, 'disability_type', '') or '',
            'نعم' if getattr(student, 'exempt_from_arabic', False) else 'لا',
            'نعم' if getattr(student, 'exempt_from_english', False) else 'لا',
            'نعم' if getattr(student, 'exempt_from_french', False) else 'لا',
            getattr(student, 'other_subject_exemptions', '') or '',
            'نعم' if getattr(student, 'is_staff_child', False) else 'لا',
            getattr(student, 'staff_parent_name', '') or '',
            getattr(student, 'staff_parent_job', '') or '',
            student.parent_name or '',
            student.parent_phone or '',
            student.parent_email or '',
            getattr(student, 'father_job', '') or '',
            student.get_educational_guardian_display() if hasattr(student, 'get_educational_guardian_display') else '',
            getattr(student, 'educational_guardian_name', '') or '',
            getattr(student, 'educational_guardian_phone', '') or '',
            float(student.total_fees or 0),
            float(student.total_payments or 0),
            float(student.total_owed or 0),
            student.created_at.strftime('%Y-%m-%d %H:%M:%S') if student.created_at else '',
            'نشط' if student.is_active else 'غير نشط',
        ]

    def export_students(self, request):
        export_format = request.GET.get('format', 'excel')
        grade_filter = request.GET.get('grade')

        queryset = Student.objects.filter(is_active=True).select_related(
            'academic_year',
            'grade_level__education_level'
        )

        if grade_filter:
            queryset = queryset.filter(grade_level_id=grade_filter)

        if export_format == 'csv':
            return self._export_csv(queryset)
        if export_format == 'excel':
            return self._export_excel(queryset)
        if export_format == 'json':
            return self._export_json(queryset)

        messages.error(request, 'صيغة التصدير غير مدعومة')
        return redirect('admin:students_student_changelist')

    def _export_csv(self, queryset):
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow(self.get_export_headers())

        for student in queryset:
            writer.writerow(self.student_to_export_row(student))

        return response

    def _export_excel(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "بيانات الطلاب"
        ws.sheet_view.rightToLeft = True

        headers = self.get_export_headers()

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_num, student in enumerate(queryset, 2):
            row_values = self.student_to_export_row(student)
            for col, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                if col in [34, 35, 36]:
                    cell.number_format = '#,##0.00'

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(headers)).coordinate}'

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 45)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    def _export_json(self, queryset):
        data = []
        for student in queryset:
            data.append({
                'name': student.name,
                'student_type': getattr(student, 'student_type', ''),
                'national_number': student.national_number,
                'passport_number': getattr(student, 'passport_number', ''),
                'nationality': getattr(student, 'nationality', ''),
                'religion': getattr(student, 'religion', ''),
                'age': student.age,
                'gender': student.gender,
                'date_of_birth': student.date_of_birth.isoformat() if student.date_of_birth else None,
                'phone_number': student.phone_number,
                'address': student.address,
                'academic_year': student.academic_year.name if student.academic_year else None,
                'education_level': student.grade_level.education_level.name if (student.grade_level and getattr(student.grade_level, 'education_level', None)) else None,
                'grade_level': student.grade_level.name if student.grade_level else None,
                'enrollment_status': getattr(student, 'enrollment_status', ''),
                'transferred_from_school': getattr(student, 'transferred_from_school', ''),
                'transferred_to_school': getattr(student, 'transferred_to_school', ''),
                'is_integration_student': getattr(student, 'is_integration_student', False),
                'disability_type': getattr(student, 'disability_type', ''),
                'subject_exemptions': student.get_subject_exemptions_display() if hasattr(student, 'get_subject_exemptions_display') else '',
                'is_staff_child': getattr(student, 'is_staff_child', False),
                'parent_name': student.parent_name,
                'parent_phone': student.parent_phone,
                'parent_email': student.parent_email,
                'father_job': getattr(student, 'father_job', ''),
                'total_fees': float(student.total_fees or 0),
                'total_payments': float(student.total_payments or 0),
                'total_owed': float(student.total_owed or 0),
                'created_at': student.created_at.isoformat() if student.created_at else None,
                'is_active': student.is_active,
            })

        response = JsonResponse({
            'students': data,
            'export_date': datetime.now().isoformat(),
            'total_count': len(data)
        }, json_dumps_params={'ensure_ascii': False, 'indent': 2})
        response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
        return response

    def download_template(self, request):
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.workbook.defined_name import DefinedName
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        def sanitize_range_name(value):
            value = str(value or '').strip()
            if not value:
                return 'EMPTY'
            value = value.replace(' ', '_').replace('-', '_').replace('/', '_').replace('\\', '_')
            value = value.replace('(', '_').replace(')', '_').replace('.', '_').replace(',', '_')
            value = value.replace('،', '_').replace(':', '_')
            while '__' in value:
                value = value.replace('__', '_')
            if value and value[0].isdigit():
                value = f'LVL_{value}'
            return value

        header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        center_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

        academic_years = list(
            SettingsAcademicYear.objects.filter(is_active=True).order_by('-start_date', '-id')
        )

        grade_levels = list(
            GradeLevel.objects.filter(is_active=True)
            .select_related('education_level')
            .order_by('education_level__order', 'order', 'name')
        )

        education_levels = []
        education_level_map = {}
        seen_education_levels = set()

        for grade_level in grade_levels:
            education_name = grade_level.education_level.name if getattr(grade_level, 'education_level', None) else ''
            if not education_name:
                continue

            if education_name not in seen_education_levels:
                seen_education_levels.add(education_name)
                education_levels.append(education_name)

            education_level_map.setdefault(education_name, [])
            education_level_map[education_name].append(grade_level.name)

        ws = wb.active
        ws.title = "قالب الاستيراد"
        ws.sheet_view.rightToLeft = True

        headers = self.get_import_headers()

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        sample_rows = [
            [
                'أحمد محمد علي',
                'طالب عادي',
                '30012010012345',
                '',
                'مصري',
                'مسلم',
                '14',
                'M',
                '2010-01-01',
                '01234567890',
                'القاهرة',
                academic_years[0].name if academic_years else '',
                education_levels[0] if education_levels else '',
                education_level_map.get(education_levels[0], [''])[0] if education_levels else '',
                'مستجد',
                '',
                '',
                'لا',
                '',
                'لا',
                'لا',
                'لا',
                '',
                'لا',
                '',
                '',
                'محمد علي أحمد',
                '01098765432',
                'parent@email.com',
                'محاسب',
                'الأب',
                '',
                '',
                '15000',
                '5000',
            ]
        ]

        for row_num, row_data in enumerate(sample_rows, 2):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = center_alignment

        instructions = [
            "تعليمات سريعة:",
            "1. استخدم نفس أسماء الأعمدة بدون تغيير.",
            "2. استخدم القوائم المنسدلة لاختيار العام الدراسي والمرحلة التعليمية والصف الدراسي.",
            "3. قائمة الصف الدراسي تعتمد على المرحلة التعليمية المختارة في نفس الصف.",
            "4. الاسم هو الحقل الإجباري الوحيد.",
            "5. الرقم القومي أو جواز السفر يجب ألا يكون مكرراً إذا تم إدخاله.",
            "6. النوع يقبل: M / F أو ذكر / أنثى.",
            "7. تاريخ الميلاد يفضل بصيغة YYYY-MM-DD.",
        ]

        start_row = len(sample_rows) + 4
        for index, instruction in enumerate(instructions):
            cell = ws.cell(row=start_row + index, column=1, value=instruction)
            if index == 0:
                cell.font = openpyxl.styles.Font(bold=True, color="D32F2F")
            else:
                cell.font = openpyxl.styles.Font(color="666666")

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(headers)).coordinate}'

        ws_years = wb.create_sheet(title="الاعوام الدراسية")
        ws_years.sheet_view.rightToLeft = True

        for col, header in enumerate(['م', 'العام الدراسي', 'نشط؟'], 1):
            cell = ws_years.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for row_num, academic_year in enumerate(academic_years, 2):
            ws_years.cell(row=row_num, column=1, value=row_num - 1)
            ws_years.cell(row=row_num, column=2, value=getattr(academic_year, 'name', str(academic_year)))
            ws_years.cell(row=row_num, column=3, value='نعم' if getattr(academic_year, 'is_active', False) else 'لا')

        ws_years.freeze_panes = 'A2'
        ws_years.auto_filter.ref = f'A1:C{max(ws_years.max_row, 2)}'

        ws_grades = wb.create_sheet(title="المراحل والصفوف")
        ws_grades.sheet_view.rightToLeft = True

        for col, header in enumerate(['م', 'المرحلة التعليمية', 'اسم النطاق', 'الصف الدراسي', 'نشط؟'], 1):
            cell = ws_grades.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        grade_row = 2
        for education_name, grades in education_level_map.items():
            range_name = sanitize_range_name(education_name)
            for grade_name in grades:
                ws_grades.cell(row=grade_row, column=1, value=grade_row - 1)
                ws_grades.cell(row=grade_row, column=2, value=education_name)
                ws_grades.cell(row=grade_row, column=3, value=range_name)
                ws_grades.cell(row=grade_row, column=4, value=grade_name)
                ws_grades.cell(row=grade_row, column=5, value='نعم')
                grade_row += 1

        ws_grades.freeze_panes = 'A2'
        ws_grades.auto_filter.ref = f'A1:E{max(ws_grades.max_row, 2)}'

        ws_choices = wb.create_sheet(title="قيم مسموح بها")
        ws_choices.sheet_view.rightToLeft = True

        for col, header in enumerate(['نوع البيان', 'القيمة المعروضة', 'القيمة الداخلية'], 1):
            cell = ws_choices.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        row_num = 2

        for value, label in getattr(Student, 'STUDENT_TYPE_CHOICES', []):
            ws_choices.cell(row=row_num, column=1, value='نوع الطالب')
            ws_choices.cell(row=row_num, column=2, value=str(label))
            ws_choices.cell(row=row_num, column=3, value=str(value))
            row_num += 1

        for value, label in getattr(Student, 'RELIGION_CHOICES', []):
            ws_choices.cell(row=row_num, column=1, value='الديانة')
            ws_choices.cell(row=row_num, column=2, value=str(label))
            ws_choices.cell(row=row_num, column=3, value=str(value))
            row_num += 1

        for value, label in getattr(Student, 'ENROLLMENT_STATUS_CHOICES', []):
            ws_choices.cell(row=row_num, column=1, value='حالة القيد')
            ws_choices.cell(row=row_num, column=2, value=str(label))
            ws_choices.cell(row=row_num, column=3, value=str(value))
            row_num += 1

        for value, label in getattr(Student, 'EDUCATIONAL_GUARDIAN_CHOICES', []):
            ws_choices.cell(row=row_num, column=1, value='صاحب الولاية التعليمية')
            ws_choices.cell(row=row_num, column=2, value=str(label))
            ws_choices.cell(row=row_num, column=3, value=str(value))
            row_num += 1

        extra_values = [
            ('النوع', 'M', 'M'),
            ('النوع', 'F', 'F'),
            ('النوع', 'ذكر', 'M'),
            ('النوع', 'أنثى', 'F'),
            ('القيم المنطقية', 'نعم', 'True'),
            ('القيم المنطقية', 'لا', 'False'),
            ('القيم المنطقية', '1', 'True'),
            ('القيم المنطقية', '0', 'False'),
        ]

        for item_type, display_value, internal_value in extra_values:
            ws_choices.cell(row=row_num, column=1, value=item_type)
            ws_choices.cell(row=row_num, column=2, value=display_value)
            ws_choices.cell(row=row_num, column=3, value=internal_value)
            row_num += 1

        ws_choices.freeze_panes = 'A2'
        ws_choices.auto_filter.ref = f'A1:C{max(ws_choices.max_row, 2)}'

        ws_lists = wb.create_sheet(title="_lists")
        ws_lists.sheet_view.rightToLeft = True

        student_type_labels = [str(label) for value, label in getattr(Student, 'STUDENT_TYPE_CHOICES', [])]
        religion_labels = [str(label) for value, label in getattr(Student, 'RELIGION_CHOICES', [])]
        enrollment_labels = [str(label) for value, label in getattr(Student, 'ENROLLMENT_STATUS_CHOICES', [])]
        guardian_labels = [str(label) for value, label in getattr(Student, 'EDUCATIONAL_GUARDIAN_CHOICES', [])]
        gender_values = ['M', 'F', 'ذكر', 'أنثى']
        yes_no_values = ['نعم', 'لا', '1', '0']

        for idx, value in enumerate(student_type_labels, 1):
            ws_lists.cell(row=idx, column=1, value=value)
        for idx, value in enumerate(religion_labels, 1):
            ws_lists.cell(row=idx, column=2, value=value)
        for idx, value in enumerate(enrollment_labels, 1):
            ws_lists.cell(row=idx, column=3, value=value)
        for idx, value in enumerate(guardian_labels, 1):
            ws_lists.cell(row=idx, column=4, value=value)
        for idx, value in enumerate(gender_values, 1):
            ws_lists.cell(row=idx, column=5, value=value)
        for idx, value in enumerate(yes_no_values, 1):
            ws_lists.cell(row=idx, column=6, value=value)
        for idx, academic_year in enumerate(academic_years, 1):
            ws_lists.cell(row=idx, column=7, value=getattr(academic_year, 'name', str(academic_year)))
        for idx, value in enumerate(education_levels, 1):
            ws_lists.cell(row=idx, column=8, value=value)

        from openpyxl.workbook.defined_name import DefinedName
        max_grade_list_len = 1
        dependent_start_col = 9

        for offset, education_name in enumerate(education_levels):
            col_num = dependent_start_col + offset
            col_letter = get_column_letter(col_num)
            range_name = sanitize_range_name(education_name)

            ws_lists.cell(row=1, column=col_num, value=education_name)
            grade_names = education_level_map.get(education_name, [])
            max_grade_list_len = max(max_grade_list_len, len(grade_names))

            for grade_idx, grade_name in enumerate(grade_names, 2):
                ws_lists.cell(row=grade_idx, column=col_num, value=grade_name)

            end_row = max(len(grade_names) + 1, 2)
            wb.defined_names.add(
                DefinedName(
                    range_name,
                    attr_text=f"'_lists'!${col_letter}$2:${col_letter}${end_row}"
                )
            )

        wb.defined_names.add(DefinedName('student_type_choices', attr_text=f"'_lists'!$A$1:$A${max(len(student_type_labels), 1)}"))
        wb.defined_names.add(DefinedName('religion_choices', attr_text=f"'_lists'!$B$1:$B${max(len(religion_labels), 1)}"))
        wb.defined_names.add(DefinedName('enrollment_choices', attr_text=f"'_lists'!$C$1:$C${max(len(enrollment_labels), 1)}"))
        wb.defined_names.add(DefinedName('guardian_choices', attr_text=f"'_lists'!$D$1:$D${max(len(guardian_labels), 1)}"))
        wb.defined_names.add(DefinedName('gender_choices', attr_text=f"'_lists'!$E$1:$E${max(len(gender_values), 1)}"))
        wb.defined_names.add(DefinedName('boolean_choices', attr_text=f"'_lists'!$F$1:$F${max(len(yes_no_values), 1)}"))
        wb.defined_names.add(DefinedName('academic_year_choices', attr_text=f"'_lists'!$G$1:$G${max(len(academic_years), 1)}"))
        wb.defined_names.add(DefinedName('education_level_choices', attr_text=f"'_lists'!$H$1:$H${max(len(education_levels), 1)}"))

        ws_lists.sheet_state = 'hidden'

        max_rows = 5000

        dv_student_type = DataValidation(type="list", formula1="=student_type_choices", allow_blank=True)
        ws.add_data_validation(dv_student_type)
        dv_student_type.add(f'B2:B{max_rows}')

        dv_religion = DataValidation(type="list", formula1="=religion_choices", allow_blank=True)
        ws.add_data_validation(dv_religion)
        dv_religion.add(f'F2:F{max_rows}')

        dv_gender = DataValidation(type="list", formula1="=gender_choices", allow_blank=True)
        ws.add_data_validation(dv_gender)
        dv_gender.add(f'H2:H{max_rows}')

        dv_academic_year = DataValidation(type="list", formula1="=academic_year_choices", allow_blank=True)
        ws.add_data_validation(dv_academic_year)
        dv_academic_year.add(f'L2:L{max_rows}')

        dv_education_level = DataValidation(type="list", formula1="=education_level_choices", allow_blank=True)
        ws.add_data_validation(dv_education_level)
        dv_education_level.add(f'M2:M{max_rows}')

        for row in range(2, max_rows + 1):
            dv_grade_level = DataValidation(
                type="list",
                formula1=f'=INDIRECT(SUBSTITUTE($M{row}," ","_"))',
                allow_blank=True
            )
            ws.add_data_validation(dv_grade_level)
            dv_grade_level.add(f'N{row}')

        dv_enrollment = DataValidation(type="list", formula1="=enrollment_choices", allow_blank=True)
        ws.add_data_validation(dv_enrollment)
        dv_enrollment.add(f'O2:O{max_rows}')

        dv_boolean = DataValidation(type="list", formula1="=boolean_choices", allow_blank=True)
        ws.add_data_validation(dv_boolean)
        for col in ['R', 'T', 'U', 'V', 'X']:
            dv_boolean.add(f'{col}2:{col}{max_rows}')

        dv_guardian = DataValidation(type="list", formula1="=guardian_choices", allow_blank=True)
        ws.add_data_validation(dv_guardian)
        dv_guardian.add(f'AE2:AE{max_rows}')

        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        value_length = len(str(cell.value)) if cell.value else 0
                        if value_length > max_length:
                            max_length = value_length
                    except Exception:
                        pass
                sheet.column_dimensions[column_letter].width = min(max_length + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        return response

    def download_reference_data(self, request):
        wb = openpyxl.Workbook()

        header_fill = openpyxl.styles.PatternFill(
            start_color="1F4E79",
            end_color="1F4E79",
            fill_type="solid"
        )
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        center_alignment = openpyxl.styles.Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        def style_header(ws, headers):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

        def autosize(ws, max_width=40):
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        value_length = len(str(cell.value)) if cell.value else 0
                        if value_length > max_length:
                            max_length = value_length
                    except Exception:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 4, max_width)

        ws_info = wb.active
        ws_info.title = "شرح الاستخدام"
        ws_info.sheet_view.rightToLeft = True

        info_rows = [
            ["ملف القيم المرجعية لاستيراد الطلاب"],
            ["هذا الملف يساعدك على معرفة القيم الصحيحة التي يجب استخدامها داخل ملف استيراد الطلاب."],
            ["استخدم نفس أسماء الأعوام الدراسية والمراحل التعليمية والصفوف كما هي مكتوبة هنا."],
            ["إذا كان اسم الصف موجوداً في أكثر من مرحلة، يجب كتابة المرحلة التعليمية أيضاً."],
            ["يمكنك تحميل القالب الذكي واختيار القيم من القوائم المنسدلة لتقليل الأخطاء."],
            ["الشيتات الموجودة في هذا الملف:"],
            ["1. الاعوام الدراسية: كل الأعوام الدراسية النشطة."],
            ["2. المراحل والصفوف: كل الصفوف الدراسية النشطة مع المرحلة التعليمية التابعة لها."],
            ["3. قيم مسموح بها: القيم المعتمدة لبعض الحقول مثل نوع الطالب والديانة وحالة القيد."],
        ]

        for row_num, row_data in enumerate(info_rows, 1):
            cell = ws_info.cell(row=row_num, column=1, value=row_data[0])
            if row_num == 1:
                cell.font = openpyxl.styles.Font(bold=True, size=14, color="1F1F1F")
            else:
                cell.font = openpyxl.styles.Font(size=12, color="444444")

        ws_info.column_dimensions["A"].width = 100

        ws_years = wb.create_sheet(title="الاعوام الدراسية")
        ws_years.sheet_view.rightToLeft = True
        style_header(ws_years, ['م', 'العام الدراسي', 'نشط؟'])

        academic_years = SettingsAcademicYear.objects.all().order_by('-start_date', '-id')
        for row_num, academic_year in enumerate(academic_years, 2):
            ws_years.cell(row=row_num, column=1, value=row_num - 1)
            ws_years.cell(row=row_num, column=2, value=getattr(academic_year, 'name', str(academic_year)))
            ws_years.cell(row=row_num, column=3, value='نعم' if getattr(academic_year, 'is_active', False) else 'لا')

        ws_years.freeze_panes = 'A2'
        ws_years.auto_filter.ref = f'A1:C{max(ws_years.max_row, 2)}'
        autosize(ws_years)

        ws_grades = wb.create_sheet(title="المراحل والصفوف")
        ws_grades.sheet_view.rightToLeft = True
        style_header(ws_grades, ['م', 'المرحلة التعليمية', 'الصف الدراسي', 'نشط؟'])

        grade_levels = GradeLevel.objects.select_related('education_level').all().order_by(
            'education_level__order',
            'order',
            'name'
        )

        for row_num, grade_level in enumerate(grade_levels, 2):
            ws_grades.cell(row=row_num, column=1, value=row_num - 1)
            ws_grades.cell(
                row=row_num,
                column=2,
                value=grade_level.education_level.name if getattr(grade_level, 'education_level', None) else ''
            )
            ws_grades.cell(row=row_num, column=3, value=grade_level.name)
            ws_grades.cell(
                row=row_num,
                column=4,
                value='نعم' if getattr(grade_level, 'is_active', False) else 'لا'
            )

        ws_grades.freeze_panes = 'A2'
        ws_grades.auto_filter.ref = f'A1:D{max(ws_grades.max_row, 2)}'
        autosize(ws_grades)

        ws_choices = wb.create_sheet(title="قيم مسموح بها")
        ws_choices.sheet_view.rightToLeft = True
        style_header(ws_choices, ['نوع البيان', 'القيمة المعروضة', 'القيمة الداخلية'])

        row_num = 2

        for value, label in getattr(Student, 'STUDENT_TYPE_CHOICES', []):
            ws_choices.cell(row=row_num, column=1, value='نوع الطالب')
            ws_choices.cell(row=row_num, column=2, value=str(label))
            ws_choices.cell(row=row_num, column=3, value=str(value))
            row_num += 1

        for value, label in getattr(Student, 'RELIGION_CHOICES', []):
            ws_choices.cell(row=row_num, column=1, value='الديانة')
            ws_choices.cell(row=row_num, column=2, value=str(label))
            ws_choices.cell(row=row_num, column=3, value=str(value))
            row_num += 1

        for value, label in getattr(Student, 'ENROLLMENT_STATUS_CHOICES', []):
            ws_choices.cell(row=row_num, column=1, value='حالة القيد')
            ws_choices.cell(row=row_num, column=2, value=str(label))
            ws_choices.cell(row=row_num, column=3, value=str(value))
            row_num += 1

        for value, label in getattr(Student, 'EDUCATIONAL_GUARDIAN_CHOICES', []):
            ws_choices.cell(row=row_num, column=1, value='صاحب الولاية التعليمية')
            ws_choices.cell(row=row_num, column=2, value=str(label))
            ws_choices.cell(row=row_num, column=3, value=str(value))
            row_num += 1

        extra_values = [
            ('النوع', 'M', 'M'),
            ('النوع', 'F', 'F'),
            ('النوع', 'ذكر', 'M'),
            ('النوع', 'أنثى', 'F'),
            ('القيم المنطقية', 'نعم', 'True'),
            ('القيم المنطقية', 'لا', 'False'),
            ('القيم المنطقية', '1', 'True'),
            ('القيم المنطقية', '0', 'False'),
        ]

        for item_type, display_value, internal_value in extra_values:
            ws_choices.cell(row=row_num, column=1, value=item_type)
            ws_choices.cell(row=row_num, column=2, value=display_value)
            ws_choices.cell(row=row_num, column=3, value=internal_value)
            row_num += 1

        ws_choices.freeze_panes = 'A2'
        ws_choices.auto_filter.ref = f'A1:C{max(ws_choices.max_row, 2)}'
        autosize(ws_choices)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="student_import_reference_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    def import_students_view(self, request):
        grade_levels = GradeLevel.objects.filter(is_active=True).select_related('education_level').order_by(
            'education_level__order', 'order'
        )
        academic_years = SettingsAcademicYear.objects.filter(is_active=True).order_by('-start_date')

        context = {
            'title': 'استيراد بيانات الطلاب',
            'grade_levels': grade_levels,
            'academic_years': academic_years,
            'opts': Student._meta,
        }
        return render(request, 'admin/students/student/import_students.html', context)

    def process_import(self, request):
        if request.method != 'POST':
            return redirect('admin:students_student_import')

        file = request.FILES.get('import_file')
        if not file:
            messages.error(request, 'يرجى اختيار ملف للاستيراد')
            return redirect('admin:students_student_import')

        try:
            if file.name.endswith('.xlsx'):
                success_count, errors = self._process_excel_import(file, request)
            elif file.name.endswith('.csv'):
                success_count, errors = self._process_csv_import(file, request)
            else:
                messages.error(request, 'صيغة الملف غير مدعومة. يرجى استخدام Excel أو CSV')
                return redirect('admin:students_student_import')

            if success_count > 0:
                messages.success(request, f'تم استيراد {success_count} طالب بنجاح')

            if errors:
                error_msg = f'حدثت {len(errors)} أخطاء أثناء الاستيراد:\n' + '\n'.join(errors[:8])
                if len(errors) > 8:
                    error_msg += f'\n... و {len(errors) - 8} أخطاء أخرى'
                messages.warning(request, error_msg)

        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء معالجة الملف: {str(e)}')

        return redirect('admin:students_student_changelist')

    def _process_excel_import(self, file, request):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        success_count = 0
        errors = []
        headers = [cell.value for cell in ws[1]]

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):
                    continue

                try:
                    student_data = dict(zip(headers, row))
                    student = self._create_student_from_data(student_data, row_num)
                    if student:
                        success_count += 1
                except ValidationError as e:
                    errors.append(f'الصف {row_num}: {", ".join(e.messages)}')
                except Exception as e:
                    errors.append(f'الصف {row_num}: {str(e)}')

        return success_count, errors

    def _process_csv_import(self, file, request):
        success_count = 0
        errors = []
        file_content = file.read().decode('utf-8-sig')
        csv_reader = csv.DictReader(io.StringIO(file_content))

        with transaction.atomic():
            for row_num, row in enumerate(csv_reader, 2):
                try:
                    student = self._create_student_from_data(row, row_num)
                    if student:
                        success_count += 1
                except ValidationError as e:
                    errors.append(f'الصف {row_num}: {", ".join(e.messages)}')
                except Exception as e:
                    errors.append(f'الصف {row_num}: {str(e)}')

        return success_count, errors

    def _create_student_from_data(self, data, row_num):
        name = get_text_value(data, 'الاسم*', 'الاسم', 'name')
        if not name:
            raise ValidationError('اسم الطالب مطلوب')

        national_number = get_text_value(data, 'الرقم القومي*', 'الرقم القومي', 'national_number')
        passport_number = get_text_value(data, 'رقم جواز السفر', 'passport_number')

        if national_number and Student.objects.filter(national_number=national_number).exists():
            raise ValidationError(f'الرقم القومي {national_number} موجود مسبقاً')

        if passport_number and Student.objects.filter(passport_number=passport_number).exists():
            raise ValidationError(f'رقم جواز السفر {passport_number} موجود مسبقاً')

        age = get_text_value(data, 'العمر', 'age')
        try:
            age = int(age) if age else None
        except (ValueError, TypeError):
            age = None

        gender_raw = get_text_value(data, 'النوع', 'النوع (M/F)', 'gender').strip()
        gender_map = {
            'M': 'M',
            'F': 'F',
            'ذكر': 'M',
            'أنثى': 'F',
            'male': 'M',
            'female': 'F',
        }
        gender = gender_map.get(gender_raw, gender_map.get(gender_raw.upper(), ''))

        date_of_birth = parse_date_value(get_text_value(data, 'تاريخ الميلاد', 'تاريخ الميلاد (YYYY-MM-DD)', 'date_of_birth'))

        academic_year = None
        year_name = get_text_value(data, 'العام الدراسي', 'academic_year')
        if year_name:
            academic_year = SettingsAcademicYear.objects.filter(name__iexact=year_name, is_active=True).first()
            if not academic_year:
                academic_year = SettingsAcademicYear.objects.filter(name__icontains=year_name, is_active=True).first()

        grade_level = None
        education_level_name = get_text_value(data, 'المرحلة التعليمية', 'education_level')
        grade_name = get_text_value(data, 'الصف الدراسي', 'grade_level')

        if grade_name and education_level_name:
            matches = GradeLevel.objects.select_related('education_level').filter(
                name__iexact=grade_name,
                education_level__name__iexact=education_level_name,
                is_active=True
            )
            grade_level = matches.first()
            if not grade_level:
                matches = GradeLevel.objects.select_related('education_level').filter(
                    name__icontains=grade_name,
                    education_level__name__icontains=education_level_name,
                    is_active=True
                )
                grade_level = matches.first()

            if not grade_level:
                raise ValidationError(
                    f'تعذر العثور على الصف الدراسي "{grade_name}" داخل المرحلة التعليمية "{education_level_name}"'
                )

        elif grade_name:
            matches = GradeLevel.objects.select_related('education_level').filter(
                name__iexact=grade_name,
                is_active=True
            )

            if matches.count() == 1:
                grade_level = matches.first()
            elif matches.count() > 1:
                raise ValidationError(
                    f'الصف الدراسي "{grade_name}" موجود في أكثر من مرحلة، يرجى تحديد المرحلة التعليمية'
                )
            else:
                matches = GradeLevel.objects.select_related('education_level').filter(
                    name__icontains=grade_name,
                    is_active=True
                )
                if matches.count() == 1:
                    grade_level = matches.first()
                elif matches.count() > 1:
                    raise ValidationError(
                        f'الصف الدراسي "{grade_name}" غير محدد بدقة، يرجى كتابة المرحلة التعليمية'
                    )

        student_type = get_choice_value(
            get_text_value(data, 'نوع الطالب', 'student_type'),
            Student.STUDENT_TYPE_CHOICES,
            default='REGULAR'
        )

        religion = get_choice_value(
            get_text_value(data, 'الديانة', 'religion'),
            Student.RELIGION_CHOICES,
            default=''
        )

        enrollment_status = get_choice_value(
            get_text_value(data, 'حالة القيد', 'enrollment_status'),
            Student.ENROLLMENT_STATUS_CHOICES,
            default='NEW'
        )

        educational_guardian = get_choice_value(
            get_text_value(data, 'صاحب الولاية التعليمية', 'educational_guardian'),
            Student.EDUCATIONAL_GUARDIAN_CHOICES,
            default='FATHER'
        )

        student = Student.objects.create(
            name=name,
            student_type=student_type,
            national_number=national_number or None,
            passport_number=passport_number or None,
            nationality=get_text_value(data, 'الجنسية', 'nationality'),
            religion=religion,
            age=age,
            gender=gender,
            date_of_birth=date_of_birth,
            phone_number=get_text_value(data, 'رقم الهاتف', 'phone_number'),
            address=get_text_value(data, 'العنوان', 'address'),
            grade_level=grade_level,
            academic_year=academic_year,
            enrollment_status=enrollment_status,
            transferred_from_school=get_text_value(data, 'محول من مدرسة', 'transferred_from_school'),
            transferred_to_school=get_text_value(data, 'محول إلى مدرسة', 'transferred_to_school'),
            is_integration_student=get_bool_value(data, 'طالب دمج', 'is_integration_student'),
            disability_type=get_text_value(data, 'نوع الإعاقة', 'disability_type'),
            exempt_from_arabic=get_bool_value(data, 'إعفاء من العربي', 'exempt_from_arabic'),
            exempt_from_english=get_bool_value(data, 'إعفاء من الإنجليزي', 'exempt_from_english'),
            exempt_from_french=get_bool_value(data, 'إعفاء من الفرنسي', 'exempt_from_french'),
            other_subject_exemptions=get_text_value(data, 'إعفاءات أخرى', 'other_subject_exemptions'),
            is_staff_child=get_bool_value(data, 'من أبناء العاملين', 'is_staff_child'),
            staff_parent_name=get_text_value(data, 'اسم الموظف', 'staff_parent_name'),
            staff_parent_job=get_text_value(data, 'وظيفة الموظف', 'staff_parent_job'),
            parent_name=get_text_value(data, 'اسم ولي الأمر', 'parent_name'),
            parent_phone=get_text_value(data, 'هاتف ولي الأمر', 'parent_phone'),
            parent_email=get_text_value(data, 'بريد ولي الأمر', 'parent_email'),
            father_job=get_text_value(data, 'وظيفة الأب', 'father_job'),
            educational_guardian=educational_guardian,
            educational_guardian_name=get_text_value(data, 'اسم صاحب الولاية التعليمية', 'educational_guardian_name'),
            educational_guardian_phone=get_text_value(data, 'هاتف صاحب الولاية التعليمية', 'educational_guardian_phone'),
            total_fees=parse_decimal_value(get_text_value(data, 'إجمالي المصروفات', 'total_fees')),
            total_payments=parse_decimal_value(get_text_value(data, 'إجمالي المدفوعات', 'total_payments')),
            is_active=True,
        )

        return student


# ===================================
# إدارة الطلاب
# ===================================

@admin.register(Student)
class StudentAdmin(StudentImportExportMixin, admin.ModelAdmin):
    change_list_template = 'admin/students/student/change_list.html'

    list_display = (
        'student_avatar',
        'name_with_id',
        'identity_display_admin',
        'student_flags_display',
        'age_with_gender',
        'academic_info',
        'guardian_info',
        'financial_status_display',
        'registration_date',
        'is_active_display',
    )

    list_filter = (
        'is_active',
        StudentTypeFilter,
        EnrollmentStatusFilter,
        GenderFilter,
        ReligionFilter,
        IntegrationStudentFilter,
        StaffChildFilter,
        FinancialStatusFilter,
        AgeRangeFilter,
        'grade_level__education_level',
        'academic_year',
        'created_at',
    )

    search_fields = (
        'name',
        'national_number',
        'passport_number',
        'phone_number',
        'parent_name',
        'parent_phone',
        'educational_guardian_name',
        'father_job',
        'staff_parent_name',
        'transferred_from_school',
        'transferred_to_school',
        'address',
    )

    ordering = ('-created_at', 'name')
    list_per_page = 25

    readonly_fields = (
        'age',
        'created_at',
        'updated_at',
        'total_owed',
        'student_detail_link',
    )

    fieldsets = (
        ('البيانات الأساسية', {
            'fields': (
                'name',
                ('student_type', 'is_active'),
                ('national_number', 'passport_number'),
                ('nationality', 'religion'),
                ('age', 'gender', 'date_of_birth'),
            ),
            'classes': ('wide',)
        }),
        ('معلومات الاتصال', {
            'fields': (
                'phone_number',
                'address',
            ),
            'classes': ('wide',)
        }),
        ('البيانات الأكاديمية وحالة القيد', {
            'fields': (
                ('academic_year', 'grade_level'),
                'enrollment_status',
                ('transferred_from_school', 'transferred_to_school'),
            ),
            'classes': ('wide',)
        }),
        ('طلاب الدمج وذوي الهمم', {
            'fields': (
                'is_integration_student',
                'disability_type',
                ('exempt_from_arabic', 'exempt_from_english', 'exempt_from_french'),
                'other_subject_exemptions',
            ),
            'classes': ('wide', 'collapse')
        }),
        ('بيانات ولي الأمر والولاية التعليمية', {
            'fields': (
                ('parent_name', 'parent_phone'),
                'parent_email',
                'father_job',
                ('educational_guardian', 'educational_guardian_name'),
                'educational_guardian_phone',
            ),
            'classes': ('wide',)
        }),
        ('أبناء العاملين', {
            'fields': (
                'is_staff_child',
                ('staff_parent_name', 'staff_parent_job'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('البيانات المالية', {
            'fields': (
                ('total_fees', 'total_payments', 'total_owed'),
            ),
            'classes': ('wide',)
        }),
        ('معلومات النظام', {
            'fields': (
                ('created_at', 'updated_at'),
                'student_detail_link',
            ),
            'classes': ('collapse',)
        }),
    )

    actions = [export_students_to_csv, archive_selected_students]

    formfield_overrides = {
        models.CharField: {'widget': TextInput(attrs={'size': '40'})},
        models.TextField: {'widget': Textarea(attrs={'rows': 3, 'cols': 40})},
    }

    def student_avatar(self, obj):
        if getattr(obj, 'is_integration_student', False):
            icon = 'fa-wheelchair text-warning'
        elif getattr(obj, 'student_type', '') == 'EXPATRIATE':
            icon = 'fa-passport text-info'
        elif obj.gender == 'M':
            icon = 'fa-male text-primary'
        elif obj.gender == 'F':
            icon = 'fa-female text-danger'
        else:
            icon = 'fa-user text-secondary'

        return format_html('<i class="fas {} fa-2x" title="{}"></i>', icon, obj.name)
    student_avatar.short_description = ''

    def name_with_id(self, obj):
        return format_html(
            '<strong><a href="{}" title="عرض تفاصيل الطالب">{}</a></strong>',
            reverse('admin:students_student_change', args=[obj.pk]),
            obj.name
        )
    name_with_id.short_description = 'اسم الطالب'
    name_with_id.admin_order_field = 'name'

    def identity_display_admin(self, obj):
        if obj.national_number and len(str(obj.national_number)) == 14:
            formatted = f"{obj.national_number[:2]} {obj.national_number[2:7]} {obj.national_number[7:14]}"
            return format_html('<code>{}</code><br><small>رقم قومي</small>', formatted)

        if getattr(obj, 'passport_number', ''):
            return format_html('<code>{}</code><br><small>جواز سفر</small>', obj.passport_number)

        return '-'
    identity_display_admin.short_description = 'الهوية'

    def student_flags_display(self, obj):
        flags = []

        if getattr(obj, 'student_type', '') == 'EXPATRIATE':
            flags.append('<span class="badge badge-info">وافد</span>')

        if getattr(obj, 'is_integration_student', False):
            flags.append('<span class="badge badge-warning">دمج</span>')

        if getattr(obj, 'is_staff_child', False):
            flags.append('<span class="badge badge-success">ابن عامل</span>')

        if getattr(obj, 'enrollment_status', ''):
            flags.append(f'<span class="badge badge-secondary">{obj.get_enrollment_status_display()}</span>')

        return format_html('<br>'.join(flags)) if flags else format_html('<span class="badge badge-light">عادي</span>')
    student_flags_display.short_description = 'تصنيف الطالب'

    def age_with_gender(self, obj):
        gender_display = 'ذكر' if obj.gender == 'M' else 'أنثى' if obj.gender == 'F' else 'غير محدد'
        age_display = f"{obj.age} سنة" if obj.age else 'غير محدد'
        gender_color = 'primary' if obj.gender == 'M' else 'danger' if obj.gender == 'F' else 'secondary'

        return format_html(
            '{}<br><span class="badge badge-{}">{}</span>',
            age_display,
            gender_color,
            gender_display
        )
    age_with_gender.short_description = 'العمر والجنس'
    age_with_gender.admin_order_field = 'age'

    def academic_info(self, obj):
        grade_info = obj.grade_level.name if obj.grade_level else 'غير محدد'
        education_info = obj.grade_level.education_level.name if (obj.grade_level and obj.grade_level.education_level) else 'غير محدد'
        year_info = str(obj.academic_year) if obj.academic_year else 'غير محدد'

        transfer_text = ''
        if getattr(obj, 'transferred_from_school', ''):
            transfer_text = f'<br><small class="text-info">من: {obj.transferred_from_school}</small>'
        elif getattr(obj, 'transferred_to_school', ''):
            transfer_text = f'<br><small class="text-danger">إلى: {obj.transferred_to_school}</small>'

        return format_html(
            '<strong>{}</strong><br><small>{}</small><br><small class="text-muted">{}</small>{}',
            grade_info,
            education_info,
            year_info,
            format_html(transfer_text)
        )
    academic_info.short_description = 'المعلومات الأكاديمية'

    def guardian_info(self, obj):
        guardian = obj.get_educational_guardian_display() if hasattr(obj, 'get_educational_guardian_display') else 'غير محدد'
        parent_name = obj.parent_name or getattr(obj, 'educational_guardian_name', '') or 'غير محدد'
        phone = obj.parent_phone or getattr(obj, 'educational_guardian_phone', '') or 'غير متوفر'

        return format_html(
            '<strong>{}</strong><br><small>{}</small><br><small><i class="fas fa-phone"></i> {}</small>',
            guardian,
            parent_name,
            phone
        )
    guardian_info.short_description = 'الولاية التعليمية'

    def financial_status_display(self, obj):
        try:
            status = obj.get_financial_status()
            color = obj.get_status_color()
        except Exception:
            status = 'غير محدد'
            color = 'secondary'

        bootstrap_color = {
            'success': 'success',
            'warning': 'warning',
            'danger': 'danger',
            'secondary': 'secondary'
        }.get(color, 'secondary')

        return format_html(
            '<span class="badge badge-{}">{}</span><br><small>مستحق: <strong>{}</strong> ج.م</small>',
            bootstrap_color,
            status,
            obj.total_owed or 0
        )
    financial_status_display.short_description = 'الحالة المالية'
    financial_status_display.admin_order_field = 'total_owed'

    def registration_date(self, obj):
        if obj.created_at:
            return format_html(
                '{}<br><small class="text-muted">{}</small>',
                obj.created_at.strftime('%Y-%m-%d'),
                obj.created_at.strftime('%H:%M')
            )
        return '-'
    registration_date.short_description = 'تاريخ التسجيل'
    registration_date.admin_order_field = 'created_at'

    def is_active_display(self, obj):
        if obj.is_active:
            return format_html('<span class="badge badge-success">نشط</span>')
        return format_html('<span class="badge badge-secondary">غير نشط</span>')
    is_active_display.short_description = 'الحالة'
    is_active_display.admin_order_field = 'is_active'

    def student_detail_link(self, obj):
        if obj.pk:
            try:
                url = reverse('students:student_detail', args=[obj.pk])
                return format_html('<a href="{}" target="_blank">عرض في الموقع</a>', url)
            except Exception:
                return "-"
        return "-"
    student_detail_link.short_description = 'رابط التفاصيل'

    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'academic_year',
            'grade_level__education_level'
        )

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        queryset = self.get_queryset(request)

        active_qs = queryset.filter(is_active=True)
        extra_context['total_students'] = active_qs.count()
        extra_context['male_students'] = active_qs.filter(gender='M').count()
        extra_context['female_students'] = active_qs.filter(gender='F').count()
        extra_context['expatriate_students'] = active_qs.filter(student_type='EXPATRIATE').count()
        extra_context['integration_students'] = active_qs.filter(is_integration_student=True).count()
        extra_context['staff_children'] = active_qs.filter(is_staff_child=True).count()
        extra_context['students_owing'] = active_qs.filter(total_owed__gt=0).count()
        extra_context['total_owed'] = active_qs.aggregate(total=Sum('total_owed'))['total'] or 0
        extra_context['show_import_export'] = True

        return super().changelist_view(request, extra_context)

    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }


# ===================================
# إدارة ملفات المستخدمين
# ===================================

@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    list_display = ('get_user_full_name', 'get_username', 'phone_number', 'address', 'get_user_status')
    list_filter = ('user__is_active', 'user__is_staff')
    search_fields = ('user__username', 'user__first_name', 'user__last_name', 'phone_number', 'address')
    readonly_fields = ('user',)

    def get_user_full_name(self, obj):
        full_name = obj.user.get_full_name()
        if full_name:
            return format_html('<strong>{}</strong>', full_name)
        return format_html('<em>{}</em>', obj.user.username)
    get_user_full_name.short_description = 'الاسم الكامل'

    def get_username(self, obj):
        return format_html('<code>{}</code>', obj.user.username)
    get_username.short_description = 'اسم المستخدم'

    def get_user_status(self, obj):
        if obj.user.is_superuser:
            return format_html('<span class="badge badge-danger">مدير عام</span>')
        if obj.user.is_staff:
            return format_html('<span class="badge badge-warning">موظف</span>')
        if obj.user.is_active:
            return format_html('<span class="badge badge-success">نشط</span>')
        return format_html('<span class="badge badge-secondary">غير نشط</span>')
    get_user_status.short_description = 'حالة المستخدم'


# ===================================
# إدارة أرشيف الطلاب
# ===================================

@admin.register(ArchiveStudent)
class ArchiveStudentAdmin(admin.ModelAdmin):
    list_display = (
        'archive_name',
        'archive_identity_display',
        'archive_type_display',
        'archive_age_gender',
        'archive_academic_info',
        'archive_financial_summary',
        'archived_date',
        'archived_reason',
    )

    list_filter = (
        'archive_gender',
        'archive_student_type',
        'archive_is_integration_student',
        'archive_is_staff_child',
        'archived_date',
        'archived_reason',
    )

    search_fields = (
        'archive_name',
        'archive_national_number',
        'archive_passport_number',
        'archive_parent_name',
        'archive_parent_phone',
        'archived_reason',
    )

    readonly_fields = ('archived_date',)
    date_hierarchy = 'archived_date'
    ordering = ('-archived_date',)

    fieldsets = (
        ('بيانات الطالب المؤرشف', {
            'fields': (
                'archive_name',
                ('archive_national_number', 'archive_passport_number'),
                ('archive_student_type', 'archive_nationality', 'archive_religion'),
                ('archive_age', 'archive_gender', 'archive_date_of_birth'),
            )
        }),
        ('البيانات الأكاديمية', {
            'fields': (
                'archive_academic_year',
                'archive_grade_level',
                'archive_education_level',
                'archive_enrollment_status',
                ('archive_transferred_from_school', 'archive_transferred_to_school'),
            )
        }),
        ('بيانات الدمج', {
            'fields': (
                'archive_is_integration_student',
                'archive_disability_type',
                'archive_subject_exemptions',
            ),
            'classes': ('collapse',)
        }),
        ('بيانات ولي الأمر والعاملين', {
            'fields': (
                ('archive_parent_name', 'archive_parent_phone'),
                'archive_father_job',
                ('archive_educational_guardian', 'archive_educational_guardian_name'),
                'archive_is_staff_child',
                ('archive_staff_parent_name', 'archive_staff_parent_job'),
            ),
            'classes': ('collapse',)
        }),
        ('البيانات المالية', {
            'fields': (
                ('archive_total_fees', 'archive_total_payments', 'archive_total_owed'),
            )
        }),
        ('تفاصيل الأرشفة', {
            'fields': (
                'archived_date',
                'archived_reason',
                'archived_by',
            )
        }),
    )

    def archive_identity_display(self, obj):
        if obj.archive_national_number:
            return format_html('<code>{}</code><br><small>رقم قومي</small>', obj.archive_national_number)
        if obj.archive_passport_number:
            return format_html('<code>{}</code><br><small>جواز سفر</small>', obj.archive_passport_number)
        return '-'
    archive_identity_display.short_description = 'الهوية'

    def archive_type_display(self, obj):
        flags = []
        if obj.archive_student_type:
            flags.append(obj.archive_student_type)
        if obj.archive_is_integration_student:
            flags.append('دمج')
        if obj.archive_is_staff_child:
            flags.append('ابن عامل')
        return ' - '.join(flags) if flags else '-'
    archive_type_display.short_description = 'التصنيف'

    def archive_age_gender(self, obj):
        gender_display = 'ذكر' if obj.archive_gender == 'M' else 'أنثى' if obj.archive_gender == 'F' else 'غير محدد'
        gender_color = 'primary' if obj.archive_gender == 'M' else 'danger' if obj.archive_gender == 'F' else 'secondary'
        return format_html(
            '{} سنة<br><span class="badge badge-{}">{}</span>',
            obj.archive_age,
            gender_color,
            gender_display
        )
    archive_age_gender.short_description = 'العمر والجنس'

    def archive_academic_info(self, obj):
        return format_html(
            '<strong>{}</strong><br><small>{}</small><br><small>{}</small>',
            obj.archive_grade_level,
            obj.archive_education_level,
            obj.archive_enrollment_status or ''
        )
    archive_academic_info.short_description = 'المعلومات الأكاديمية'

    def archive_financial_summary(self, obj):
        return format_html(
            'مدفوع: <strong>{}</strong> ج.م<br>مستحق: <strong>{}</strong> ج.م',
            obj.archive_total_payments,
            obj.archive_total_owed
        )
    archive_financial_summary.short_description = 'الملخص المالي'

    def has_add_permission(self, request):
        return False

# from django.contrib import admin
# from django.utils.translation import gettext_lazy as _
# from django.urls import reverse, path
# from django.utils.html import format_html
# from django.db import models, transaction
# from django.forms import TextInput, Textarea
# from django.utils.safestring import mark_safe
# from django.contrib.admin import SimpleListFilter
# from django.db.models import Count, Sum, Q
# from django.http import HttpResponse, JsonResponse
# from django.shortcuts import render, redirect
# from django.contrib import messages
# from django.core.exceptions import ValidationError
# import csv
# import io
# import openpyxl
# from datetime import datetime
# import json

# # استيراد النماذج
# from .models import Student, UserProfile, ArchiveStudent
# from school_settings.models import AcademicYear as SettingsAcademicYear, EducationLevel, GradeLevel

# # تخصيص رأس لوحة الإدارة
# admin.site.site_header = "نظام إدارة مدرسة المنار"
# admin.site.site_title = "إدارة مدرسة المنار"
# admin.site.index_title = _('لوحة التحكم الرئيسية')

# # مرشحات مخصصة
# class GenderFilter(SimpleListFilter):
#     title = _('الجنس')
#     parameter_name = 'gender'

#     def lookups(self, request, model_admin):
#         return (
#             ('M', _('ذكر')),
#             ('F', _('أنثى')),
#         )

#     def queryset(self, request, queryset):
#         if self.value() == 'M':
#             return queryset.filter(gender='M')
#         if self.value() == 'F':
#             return queryset.filter(gender='F')

# class FinancialStatusFilter(SimpleListFilter):
#     title = _('الحالة المالية')
#     parameter_name = 'financial_status'

#     def lookups(self, request, model_admin):
#         return (
#             ('paid', _('مسدد بالكامل')),
#             ('partial', _('مسدد جزئياً')),
#             ('owing', _('مستحق عليه')),
#         )

#     def queryset(self, request, queryset):
#         if self.value() == 'paid':
#             return queryset.filter(total_owed__lte=0)
#         elif self.value() == 'partial':
#             return queryset.filter(total_owed__gt=0, total_owed__lt=models.F('total_fees'))
#         elif self.value() == 'owing':
#             return queryset.filter(total_owed__gte=models.F('total_fees'))

# class AgeRangeFilter(SimpleListFilter):
#     title = _('الفئة العمرية')
#     parameter_name = 'age_range'

#     def lookups(self, request, model_admin):
#         return (
#             ('3-6', _('3-6 سنوات (حضانة)')),
#             ('6-12', _('6-12 سنة (ابتدائي)')),
#             ('12-15', _('12-15 سنة (إعدادي)')),
#             ('15-18', _('15-18 سنة (ثانوي)')),
#         )

#     def queryset(self, request, queryset):
#         if self.value() == '3-6':
#             return queryset.filter(age__gte=3, age__lte=6)
#         elif self.value() == '6-12':
#             return queryset.filter(age__gte=6, age__lte=12)
#         elif self.value() == '12-15':
#             return queryset.filter(age__gte=12, age__lte=15)
#         elif self.value() == '15-18':
#             return queryset.filter(age__gte=15, age__lte=18)

# # إجراءات مخصصة
# def export_students_to_csv(modeladmin, request, queryset):
#     """تصدير الطلاب المحددين إلى ملف CSV"""
#     response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
#     response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
#     response.write('\ufeff')  # BOM for Arabic support
    
#     writer = csv.writer(response)
#     # كتابة الرؤوس
#     writer.writerow([
#         'اسم الطالب', 'الرقم القومي', 'العمر', 'الجنس', 'تاريخ الميلاد',
#         'رقم الهاتف', 'العنوان', 'الصف الدراسي', 'المرحلة التعليمية',
#         'إجمالي المصروفات', 'إجمالي المدفوعات', 'المستحقات',
#         'اسم ولي الأمر', 'هاتف ولي الأمر', 'تاريخ التسجيل'
#     ])
    
#     # كتابة البيانات
#     for student in queryset:
#         writer.writerow([
#             student.name,
#             student.national_number,
#             student.age,
#             'ذكر' if student.gender == 'M' else 'أنثى' if student.gender == 'F' else '',
#             student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
#             student.phone_number,
#             student.address,
#             student.grade_name if hasattr(student, 'grade_name') else '',
#             student.education_level_name if hasattr(student, 'education_level_name') else '',
#             float(student.total_fees or 0),
#             float(student.total_payments or 0),
#             float(student.total_owed or 0),
#             student.parent_name,
#             student.parent_phone,
#             student.created_at.strftime('%Y-%m-%d %H:%M') if student.created_at else ''
#         ])
    
#     return response

# export_students_to_csv.short_description = "تصدير الطلاب المحددين إلى CSV"

# def archive_selected_students(modeladmin, request, queryset):
#     """أرشفة الطلاب المحددين"""
#     archived_count = 0
#     for student in queryset:
#         # إنشاء نسخة في الأرشيف
#         ArchiveStudent.objects.create(
#             archive_name=student.name,
#             archive_national_number=student.national_number,
#             archive_age=student.age or 0,
#             archive_gender=student.gender,
#             archive_date_of_birth=student.date_of_birth,
#             archive_academic_year=str(student.academic_year) if student.academic_year else "غير محدد",
#             archive_grade_level=student.grade_name if hasattr(student, 'grade_name') else "غير محدد",
#             archive_education_level=student.education_level_name if hasattr(student, 'education_level_name') else "غير محدد",
#             archive_total_payments=student.total_payments or 0,
#             archive_total_fees=student.total_fees or 0,
#             archive_total_owed=student.total_owed or 0,
#             archived_reason='أرشفة من لوحة الإدارة'
#         )
#         archived_count += 1
    
#     # حذف الطلاب الأصليين
#     deleted_count = queryset.count()
#     queryset.delete()
    
#     modeladmin.message_user(request, f'تم أرشفة {archived_count} طالب وحذفهم من النظام الحالي.')

# archive_selected_students.short_description = "أرشفة الطلاب المحددين"

# # Mixin للاستيراد والتصدير
# class StudentImportExportMixin:
#     """Mixin لإضافة وظائف الاستيراد والتصدير للطلاب"""
    
#     def get_urls(self):
#         urls = super().get_urls()
#         custom_urls = [
#             path('export/', self.admin_site.admin_view(self.export_students), name='students_student_export'),
#             path('import/', self.admin_site.admin_view(self.import_students_view), name='students_student_import'),
#             path('import/process/', self.admin_site.admin_view(self.process_import), name='students_student_import_process'),
#             path('export/template/', self.admin_site.admin_view(self.download_template), name='students_student_export_template'),
#         ]
#         return custom_urls + urls
    
#     def export_students(self, request):
#         """تصدير بيانات الطلاب"""
#         export_format = request.GET.get('format', 'excel')
#         grade_filter = request.GET.get('grade')
        
#         # فلترة الطلاب
#         queryset = Student.objects.filter(is_active=True).select_related(
#             'academic_year',
#             'grade_level__education_level'
#         )
        
#         if grade_filter:
#             queryset = queryset.filter(grade_level_id=grade_filter)
        
#         if export_format == 'csv':
#             return self._export_csv(queryset)
#         elif export_format == 'excel':
#             return self._export_excel(queryset)
#         elif export_format == 'json':
#             return self._export_json(queryset)
#         else:
#             messages.error(request, 'صيغة التصدير غير مدعومة')
#             return redirect('admin:students_student_changelist')
    
#     def _export_csv(self, queryset):
#         """تصدير CSV"""
#         response = HttpResponse(content_type='text/csv; charset=utf-8')
#         response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
#         response.write('\ufeff')  # BOM للدعم العربي
        
#         writer = csv.writer(response)
        
#         # رأس الجدول
#         headers = [
#             'الاسم*', 'الرقم القومي*', 'العمر', 'النوع (M/F)', 'تاريخ الميلاد (YYYY-MM-DD)',
#             'رقم الهاتف', 'العنوان', 'العام الدراسي', 'الصف الدراسي',
#             'اسم ولي الأمر', 'هاتف ولي الأمر', 'بريد ولي الأمر',
#             'إجمالي المصروفات', 'إجمالي المدفوعات', 'المستحقات',
#             'تاريخ الإنشاء', 'الحالة'
#         ]
#         writer.writerow(headers)
        
#         # البيانات
#         for student in queryset:
#             row = [
#                 student.name,
#                 student.national_number,
#                 student.age or '',
#                 student.gender if student.gender else '',
#                 student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
#                 student.phone_number or '',
#                 student.address or '',
#                 student.academic_year.name if student.academic_year else '',
#                 student.grade_level.name if student.grade_level else '',
#                 student.parent_name or '',
#                 student.parent_phone or '',
#                 student.parent_email or '',
#                 float(student.total_fees or 0),
#                 float(student.total_payments or 0),
#                 float(student.total_owed or 0),
#                 student.created_at.strftime('%Y-%m-%d %H:%M:%S') if student.created_at else '',
#                 'نشط' if student.is_active else 'غير نشط'
#             ]
#             writer.writerow(row)
        
#         return response
    
#     def _export_excel(self, queryset):
#         """تصدير Excel"""
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "بيانات الطلاب"
        
#         # تنسيق الرؤوس
#         headers = [
#             'الاسم*', 'الرقم القومي*', 'العمر', 'النوع (M/F)', 'تاريخ الميلاد',
#             'رقم الهاتف', 'العنوان', 'العام الدراسي', 'الصف الدراسي',
#             'اسم ولي الأمر', 'هاتف ولي الأمر', 'بريد ولي الأمر',
#             'إجمالي المصروفات', 'إجمالي المدفوعات', 'المستحقات',
#             'تاريخ الإنشاء', 'الحالة'
#         ]
        
#         # إضافة الرؤوس مع التنسيق
#         for col, header in enumerate(headers, 1):
#             cell = ws.cell(row=1, column=col, value=header)
#             cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
#             cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#             cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        
#         # البيانات
#         for row_num, student in enumerate(queryset, 2):
#             data = [
#                 student.name,
#                 student.national_number,
#                 student.age or '',
#                 student.gender or '',
#                 student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
#                 student.phone_number or '',
#                 student.address or '',
#                 student.academic_year.name if student.academic_year else '',
#                 student.grade_level.name if student.grade_level else '',
#                 student.parent_name or '',
#                 student.parent_phone or '',
#                 student.parent_email or '',
#                 float(student.total_fees or 0),
#                 float(student.total_payments or 0),
#                 float(student.total_owed or 0),
#                 student.created_at.strftime('%Y-%m-%d %H:%M:%S') if student.created_at else '',
#                 'نشط' if student.is_active else 'غير نشط'
#             ]
            
#             for col, value in enumerate(data, 1):
#                 cell = ws.cell(row=row_num, column=col, value=value)
#                 if col in [13, 14, 15]:  # الأعمدة المالية
#                     cell.number_format = '#,##0.00'
        
#         # ضبط عرض الأعمدة
#         for column in ws.columns:
#             max_length = 0
#             column_letter = column[0].column_letter
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#             adjusted_width = min(max_length + 2, 50)
#             ws.column_dimensions[column_letter].width = adjusted_width
        
#         # حفظ الملف
#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)
        
#         response = HttpResponse(
#             output.read(),
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
#         return response
    
#     def _export_json(self, queryset):
#         """تصدير JSON"""
#         data = []
#         for student in queryset:
#             student_data = {
#                 'name': student.name,
#                 'national_number': student.national_number,
#                 'age': student.age,
#                 'gender': student.gender,
#                 'date_of_birth': student.date_of_birth.isoformat() if student.date_of_birth else None,
#                 'phone_number': student.phone_number,
#                 'address': student.address,
#                 'academic_year': student.academic_year.name if student.academic_year else None,
#                 'grade_level': student.grade_level.name if student.grade_level else None,
#                 'parent_name': student.parent_name,
#                 'parent_phone': student.parent_phone,
#                 'parent_email': student.parent_email,
#                 'total_fees': float(student.total_fees or 0),
#                 'total_payments': float(student.total_payments or 0),
#                 'total_owed': float(student.total_owed or 0),
#                 'created_at': student.created_at.isoformat() if student.created_at else None,
#                 'is_active': student.is_active,
#             }
#             data.append(student_data)
        
#         response = JsonResponse({
#             'students': data,
#             'export_date': datetime.now().isoformat(),
#             'total_count': len(data)
#         }, json_dumps_params={'ensure_ascii': False, 'indent': 2})
        
#         response['Content-Disposition'] = f'attachment; filename="students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
        
#         return response
    
#     def download_template(self, request):
#         """تحميل قالب الاستيراد"""
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "قالب استيراد الطلاب"
        
#         # الرؤوس المطلوبة
#         headers = [
#             'الاسم*', 'الرقم القومي*', 'العمر', 'النوع (M/F)', 'تاريخ الميلاد (YYYY-MM-DD)',
#             'رقم الهاتف', 'العنوان', 'العام الدراسي', 'الصف الدراسي',
#             'اسم ولي الأمر', 'هاتف ولي الأمر', 'بريد ولي الأمر'
#         ]
        
#         for col, header in enumerate(headers, 1):
#             cell = ws.cell(row=1, column=col, value=header)
#             cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
#             cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#             cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        
#         # إضافة بيانات تجريبية
#         sample_data = [
#             'أحمد محمد علي', '30012010012345', '14', 'M', '2010-01-01',
#             '01234567890', 'القاهرة، مصر الجديدة', '2024-2025', 'الثاني الإعدادي',
#             'محمد علي أحمد', '01098765432', 'ahmed.parent@email.com'
#         ]
        
#         for col, value in enumerate(sample_data, 1):
#             ws.cell(row=2, column=col, value=value)
        
#         # إضافة تعليمات
#         instructions_row = 4
#         instructions = [
#             "تعليمات الاستيراد:",
#             "1. الحقول المميزة بـ * مطلوبة",
#             "2. النوع: M للذكر، F للأنثى",
#             "3. تاريخ الميلاد بصيغة: YYYY-MM-DD",
#             "4. الرقم القومي يجب أن يكون 14 رقم",
#             "5. احذف هذه التعليمات قبل الاستيراد",
#             "6. يمكن ترك الحقول الاختيارية فارغة"
#         ]
        
#         for i, instruction in enumerate(instructions):
#             cell = ws.cell(row=instructions_row + i, column=1, value=instruction)
#             if i == 0:
#                 cell.font = openpyxl.styles.Font(bold=True, color="D32F2F")
#             else:
#                 cell.font = openpyxl.styles.Font(color="666666")
        
#         # ضبط عرض الأعمدة
#         for column in ws.columns:
#             max_length = 0
#             column_letter = column[0].column_letter
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#             adjusted_width = min(max_length + 2, 30)
#             ws.column_dimensions[column_letter].width = adjusted_width
        
#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)
        
#         response = HttpResponse(
#             output.read(),
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        
#         return response
    
#     def import_students_view(self, request):
#         """صفحة استيراد الطلاب"""
#         grade_levels = GradeLevel.objects.filter(is_active=True).select_related('education_level').order_by('education_level__order', 'order')
#         academic_years = SettingsAcademicYear.objects.filter(is_active=True).order_by('-start_date')
        
#         context = {
#             'title': 'استيراد بيانات الطلاب',
#             'grade_levels': grade_levels,
#             'academic_years': academic_years,
#             'opts': Student._meta,
#         }
#         return render(request, 'admin/students/student/import_students.html', context)
    
#     def process_import(self, request):
#         """معالجة ملف الاستيراد"""
#         if request.method != 'POST':
#             return redirect('admin:students_student_import')
        
#         file = request.FILES.get('import_file')
#         if not file:
#             messages.error(request, 'يرجى اختيار ملف للاستيراد')
#             return redirect('admin:students_student_import')
        
#         try:
#             if file.name.endswith('.xlsx'):
#                 success_count, errors = self._process_excel_import(file, request)
#             elif file.name.endswith('.csv'):
#                 success_count, errors = self._process_csv_import(file, request)
#             else:
#                 messages.error(request, 'صيغة الملف غير مدعومة. يرجى استخدام Excel أو CSV')
#                 return redirect('admin:students_student_import')
            
#             if success_count > 0:
#                 messages.success(request, f'تم استيراد {success_count} طالب بنجاح')
            
#             if errors:
#                 error_msg = f'حدثت {len(errors)} أخطاء أثناء الاستيراد:\n' + '\n'.join(errors[:5])
#                 if len(errors) > 5:
#                     error_msg += f'\n... و {len(errors) - 5} أخطاء أخرى'
#                 messages.warning(request, error_msg)
            
#         except Exception as e:
#             messages.error(request, f'حدث خطأ أثناء معالجة الملف: {str(e)}')
        
#         return redirect('admin:students_student_changelist')
    
#     def _process_excel_import(self, file, request):
#         """معالجة ملف Excel"""
#         wb = openpyxl.load_workbook(file)
#         ws = wb.active
        
#         success_count = 0
#         errors = []
        
#         # الحصول على الرؤوس من الصف الأول
#         headers = [cell.value for cell in ws[1]]
        
#         with transaction.atomic():
#             for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
#                 if not any(row):  # تخطي الصفوف الفارغة
#                     continue
                
#                 try:
#                     student_data = dict(zip(headers, row))
#                     student = self._create_student_from_data(student_data, row_num)
#                     if student:
#                         success_count += 1
                
#                 except ValidationError as e:
#                     errors.append(f'الصف {row_num}: {", ".join(e.messages)}')
#                 except Exception as e:
#                     errors.append(f'الصف {row_num}: {str(e)}')
        
#         return success_count, errors
    
#     def _process_csv_import(self, file, request):
#         """معالجة ملف CSV"""
#         success_count = 0
#         errors = []
        
#         # قراءة الملف مع دعم UTF-8
#         file_content = file.read().decode('utf-8-sig')
#         csv_reader = csv.DictReader(io.StringIO(file_content))
        
#         with transaction.atomic():
#             for row_num, row in enumerate(csv_reader, 2):
#                 try:
#                     student = self._create_student_from_data(row, row_num)
#                     if student:
#                         success_count += 1
                
#                 except ValidationError as e:
#                     errors.append(f'الصف {row_num}: {", ".join(e.messages)}')
#                 except Exception as e:
#                     errors.append(f'الصف {row_num}: {str(e)}')
        
#         return success_count, errors
    
#     def _create_student_from_data(self, data, row_num):
#         """إنشاء طالب من البيانات المستوردة"""
#         from datetime import datetime
        
#         # التحقق من البيانات المطلوبة
#         name = str(data.get('الاسم*', '') or data.get('الاسم', '') or data.get('name', '')).strip()
#         national_number = str(data.get('الرقم القومي*', '') or data.get('الرقم القومي', '') or data.get('national_number', '')).strip()
        
#         if not name or not national_number:
#             raise ValidationError('الاسم والرقم القومي مطلوبان')
        
#         # التحقق من عدم تكرار الرقم القومي
#         if Student.objects.filter(national_number=national_number).exists():
#             raise ValidationError(f'الرقم القومي {national_number} موجود مسبقاً')
        
#         # معالجة البيانات الاختيارية
#         age = data.get('العمر', '') or data.get('age', '')
#         if age:
#             try:
#                 age = int(age)
#             except (ValueError, TypeError):
#                 age = None
        
#         gender = str(data.get('النوع (M/F)', '') or data.get('النوع', '') or data.get('gender', '')).strip().upper()
#         if gender and gender not in ['M', 'F']:
#             gender = None
        
#         # معالجة تاريخ الميلاد
#         date_of_birth = None
#         birth_date_str = str(data.get('تاريخ الميلاد (YYYY-MM-DD)', '') or data.get('تاريخ الميلاد', '') or data.get('date_of_birth', '')).strip()
#         if birth_date_str:
#             try:
#                 date_of_birth = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
#             except ValueError:
#                 try:
#                     date_of_birth = datetime.strptime(birth_date_str, '%d/%m/%Y').date()
#                 except ValueError:
#                     pass
        
#         # معالجة الصف والعام الدراسي
#         grade_level = None
#         grade_name = str(data.get('الصف الدراسي', '') or data.get('grade_level', '')).strip()
#         if grade_name:
#             try:
#                 grade_level = GradeLevel.objects.filter(name__icontains=grade_name, is_active=True).first()
#             except:
#                 pass
        
#         academic_year = None
#         year_name = str(data.get('العام الدراسي', '') or data.get('academic_year', '')).strip()
#         if year_name:
#             try:
#                 academic_year = SettingsAcademicYear.objects.filter(name__icontains=year_name, is_active=True).first()
#             except:
#                 pass
        
#         # إنشاء الطالب
#         student = Student.objects.create(
#             name=name,
#             national_number=national_number,
#             age=age,
#             gender=gender,
#             date_of_birth=date_of_birth,
#             phone_number=str(data.get('رقم الهاتف', '') or data.get('phone_number', '')).strip(),
#             address=str(data.get('العنوان', '') or data.get('address', '')).strip(),
#             grade_level=grade_level,
#             academic_year=academic_year,
#             parent_name=str(data.get('اسم ولي الأمر', '') or data.get('parent_name', '')).strip(),
#             parent_phone=str(data.get('هاتف ولي الأمر', '') or data.get('parent_phone', '')).strip(),
#             parent_email=str(data.get('بريد ولي الأمر', '') or data.get('parent_email', '')).strip(),
#             is_active=True
#         )
        
#         return student

# # إدارة الطلاب المتقدمة
# @admin.register(Student)
# class StudentAdmin(StudentImportExportMixin, admin.ModelAdmin):
#     # العرض الأساسي
#     list_display = (
#         'student_avatar',
#         'name_with_id', 
#         'national_number_formatted',
#         'age_with_gender', 
#         'academic_info',
#         'financial_status_display',
#         'contact_info',
#         'registration_date',
#         'is_active_display'
#     )
    
#     # الفلاتر
#     list_filter = (
#         'is_active',
#         GenderFilter,
#         FinancialStatusFilter,
#         AgeRangeFilter,
#         'grade_level__education_level',
#         'academic_year',
#         'created_at',
#     )
    
#     # البحث
#     search_fields = (
#         'name', 
#         'national_number', 
#         'phone_number',
#         'parent_name',
#         'parent_phone',
#         'address'
#     )
    
#     # الترتيب والترقيم
#     ordering = ('-created_at', 'name')
#     list_per_page = 25
    
#     # الحقول القابلة للقراءة فقط
#     readonly_fields = (
#         'age', 
#         'date_of_birth', 
#         'gender',
#         'total_payments', 
#         'total_fees', 
#         'total_owed',
#         'created_at', 
#         'updated_at',
#         'student_detail_link'
#     )
    
#     # تنظيم الحقول في النموذج
#     fieldsets = (
#         ('البيانات الأساسية', {
#             'fields': (
#                 'name', 
#                 'national_number',
#                 ('age', 'gender', 'date_of_birth'),
#                 'is_active'
#             ),
#             'classes': ('wide',)
#         }),
        
#         ('معلومات الاتصال', {
#             'fields': (
#                 'phone_number',
#                 'address'
#             ),
#             'classes': ('wide',)
#         }),
        
#         ('البيانات الأكاديمية', {
#             'fields': (
#                 'academic_year',
#                 'grade_level'
#             ),
#             'classes': ('wide',)
#         }),
        
#         ('بيانات ولي الأمر', {
#             'fields': (
#                 'parent_name',
#                 'parent_phone', 
#                 'parent_email'
#             ),
#             'classes': ('wide', 'collapse')
#         }),
        
#         ('البيانات المالية', {
#             'fields': (
#                 ('total_fees', 'total_payments', 'total_owed'),
#             ),
#             'classes': ('wide',)
#         }),
        
#         ('معلومات النظام', {
#             'fields': (
#                 ('created_at', 'updated_at'),
#                 'student_detail_link'
#             ),
#             'classes': ('collapse',)
#         }),
#     )
    
#     # الإجراءات
#     actions = [export_students_to_csv, archive_selected_students]
    
#     # تخصيص النماذج
#     formfield_overrides = {
#         models.CharField: {'widget': TextInput(attrs={'size': '40'})},
#         models.TextField: {'widget': Textarea(attrs={'rows': 3, 'cols': 40})},
#     }

#     def student_avatar(self, obj):
#         """أيقونة الطالب"""
#         gender_icon = 'fa-male text-primary' if obj.gender == 'M' else 'fa-female text-danger' if obj.gender == 'F' else 'fa-user text-secondary'
#         return format_html(
#             '<i class="fas {} fa-2x" title="{}"></i>',
#             gender_icon,
#             obj.name
#         )
#     student_avatar.short_description = ''

#     def name_with_id(self, obj):
#         """الاسم مع رابط للتفاصيل"""
#         return format_html(
#             '<strong><a href="{}" title="عرض تفاصيل الطالب">{}</a></strong>',
#             reverse('admin:students_student_change', args=[obj.pk]),
#             obj.name
#         )
#     name_with_id.short_description = 'اسم الطالب'
#     name_with_id.admin_order_field = 'name'

#     def national_number_formatted(self, obj):
#         """الرقم القومي منسق"""
#         if obj.national_number and len(obj.national_number) == 14:
#             formatted = f"{obj.national_number[:2]} {obj.national_number[2:7]} {obj.national_number[7:14]}"
#             return format_html('<code>{}</code>', formatted)
#         return obj.national_number or '-'
#     national_number_formatted.short_description = 'الرقم القومي'
#     national_number_formatted.admin_order_field = 'national_number'

#     def age_with_gender(self, obj):
#         """العمر والجنس"""
#         gender_display = 'ذكر' if obj.gender == 'M' else 'أنثى' if obj.gender == 'F' else 'غير محدد'
#         age_display = f"{obj.age} سنة" if obj.age else 'غير محدد'
        
#         gender_color = 'primary' if obj.gender == 'M' else 'danger' if obj.gender == 'F' else 'secondary'
        
#         return format_html(
#             '{}<br><span class="badge badge-{}">{}</span>',
#             age_display,
#             gender_color,
#             gender_display
#         )
#     age_with_gender.short_description = 'العمر والجنس'
#     age_with_gender.admin_order_field = 'age'

#     def academic_info(self, obj):
#         """المعلومات الأكاديمية"""
#         grade_info = obj.grade_level.name if obj.grade_level else 'غير محدد'
#         education_info = obj.grade_level.education_level.name if (obj.grade_level and obj.grade_level.education_level) else 'غير محدد'
#         year_info = str(obj.academic_year) if obj.academic_year else 'غير محدد'
        
#         return format_html(
#             '<strong>{}</strong><br><small>{}</small><br><small class="text-muted">{}</small>',
#             grade_info,
#             education_info,
#             year_info
#         )
#     academic_info.short_description = 'المعلومات الأكاديمية'

#     def financial_status_display(self, obj):
#         """عرض الحالة المالية"""
#         try:
#             status = obj.get_financial_status()
#             color = obj.get_status_color()
#         except:
#             status = 'غير محدد'
#             color = 'secondary'
        
#         # تحديد لون Bootstrap
#         bootstrap_colors = {
#             'success': 'success',
#             'warning': 'warning',
#             'danger': 'danger',
#             'secondary': 'secondary'
#         }
#         bootstrap_color = bootstrap_colors.get(color, 'secondary')
        
#         return format_html(
#             '<span class="badge badge-{}">{}</span><br>'
#             '<small>مستحق: <strong>{}</strong> ج.م</small>',
#             bootstrap_color,
#             status,
#             obj.total_owed or 0
#         )
#     financial_status_display.short_description = 'الحالة المالية'
#     financial_status_display.admin_order_field = 'total_owed'

#     def contact_info(self, obj):
#         """معلومات الاتصال"""
#         phone = obj.phone_number or obj.parent_phone or 'غير متوفر'
#         parent_name = obj.parent_name or 'غير محدد'
        
#         return format_html(
#             '<i class="fas fa-phone"></i> {}<br>'
#             '<i class="fas fa-user"></i> <small>{}</small>',
#             phone,
#             parent_name
#         )
#     contact_info.short_description = 'معلومات الاتصال'

#     def registration_date(self, obj):
#         """تاريخ التسجيل"""
#         if obj.created_at:
#             return format_html(
#                 '{}<br><small class="text-muted">{}</small>',
#                 obj.created_at.strftime('%Y-%m-%d'),
#                 obj.created_at.strftime('%H:%M')
#             )
#         return '-'
#     registration_date.short_description = 'تاريخ التسجيل'
#     registration_date.admin_order_field = 'created_at'

#     def is_active_display(self, obj):
#         """حالة النشاط مع رابط للتغيير"""
#         if obj.is_active:
#             status_badge = '<span class="badge badge-success">نشط</span>'
#         else:
#             status_badge = '<span class="badge badge-secondary">غير نشط</span>'
        
#         toggle_url = reverse('admin:students_student_change', args=[obj.pk])
#         return format_html(
#             '{}<br><small><a href="{}">تغيير الحالة</a></small>',
#             status_badge,
#             toggle_url
#         )
#     is_active_display.short_description = 'الحالة'
#     is_active_display.admin_order_field = 'is_active'

#     def student_detail_link(self, obj):
#         """رابط تفاصيل الطالب"""
#         if obj.pk:
#             try:
#                 url = reverse('students:student_detail', args=[obj.pk])
#                 return format_html('<a href="{}" target="_blank">عرض في الموقع</a>', url)
#             except:
#                 return "-"
#         return "-"
#     student_detail_link.short_description = 'رابط التفاصيل'

#     def get_queryset(self, request):
#         """تحسين الاستعلامات"""
#         return super().get_queryset(request).select_related(
#             'academic_year',
#             'grade_level__education_level'
#         ).prefetch_related(
#             'grade_level'
#         )

#     def changelist_view(self, request, extra_context=None):
#         """إضافة إحصائيات للصفحة الرئيسية وتفعيل الاستيراد/التصدير"""
#         extra_context = extra_context or {}
        
#         # إحصائيات سريعة
#         queryset = self.get_queryset(request)
#         extra_context['total_students'] = queryset.filter(is_active=True).count()
#         extra_context['male_students'] = queryset.filter(gender='M', is_active=True).count()
#         extra_context['female_students'] = queryset.filter(gender='F', is_active=True).count()
#         extra_context['students_owing'] = queryset.filter(total_owed__gt=0, is_active=True).count()
#         extra_context['total_owed'] = queryset.filter(is_active=True).aggregate(
#             total=Sum('total_owed')
#         )['total'] or 0
        
#         # تفعيل أزرار الاستيراد والتصدير
#         extra_context['show_import_export'] = True
        
#         return super().changelist_view(request, extra_context)

#     class Media:
#         css = {
#             'all': ('admin/css/custom_admin.css',)
#         }

# # باقي فئات الإدارة
# @admin.register(UserProfile)
# class UserProfileAdmin(admin.ModelAdmin):
#     list_display = ('get_user_full_name', 'get_username', 'phone_number', 'address', 'get_user_status')
#     list_filter = ('user__is_active', 'user__is_staff')
#     search_fields = ('user__username', 'user__first_name', 'user__last_name', 'phone_number', 'address')
#     readonly_fields = ('user',)

#     def get_user_full_name(self, obj):
#         full_name = obj.user.get_full_name()
#         if full_name:
#             return format_html('<strong>{}</strong>', full_name)
#         return format_html('<em>{}</em>', obj.user.username)
#     get_user_full_name.short_description = 'الاسم الكامل'

#     def get_username(self, obj):
#         return format_html('<code>{}</code>', obj.user.username)
#     get_username.short_description = 'اسم المستخدم'

#     def get_user_status(self, obj):
#         if obj.user.is_superuser:
#             return format_html('<span class="badge badge-danger">مدير عام</span>')
#         elif obj.user.is_staff:
#             return format_html('<span class="badge badge-warning">موظف</span>')
#         elif obj.user.is_active:
#             return format_html('<span class="badge badge-success">نشط</span>')
#         else:
#             return format_html('<span class="badge badge-secondary">غير نشط</span>')
#     get_user_status.short_description = 'حالة المستخدم'


# @admin.register(ArchiveStudent)
# class ArchiveStudentAdmin(admin.ModelAdmin):
#     list_display = (
#         'archive_name', 
#         'archive_national_number', 
#         'archive_age_gender',
#         'archive_academic_info',
#         'archive_financial_summary',
#         'archived_date',
#         'archived_reason'
#     )
#     list_filter = (
#         'archive_gender', 
#         'archived_date',
#         'archived_reason'
#     )
#     search_fields = ('archive_name', 'archive_national_number', 'archived_reason')
#     readonly_fields = ('archived_date',)
#     date_hierarchy = 'archived_date'
#     ordering = ('-archived_date',)

#     def archive_age_gender(self, obj):
#         gender_display = 'ذكر' if obj.archive_gender == 'M' else 'أنثى' if obj.archive_gender == 'F' else 'غير محدد'
#         gender_color = 'primary' if obj.archive_gender == 'M' else 'danger' if obj.archive_gender == 'F' else 'secondary'
        
#         return format_html(
#             '{} سنة<br><span class="badge badge-{}">{}</span>',
#             obj.archive_age,
#             gender_color,
#             gender_display
#         )
#     archive_age_gender.short_description = 'العمر والجنس'

#     def archive_academic_info(self, obj):
#         return format_html(
#             '<strong>{}</strong><br><small>{}</small>',
#             obj.archive_grade_level,
#             obj.archive_education_level
#         )
#     archive_academic_info.short_description = 'المعلومات الأكاديمية'

#     def archive_financial_summary(self, obj):
#         return format_html(
#             'مدفوع: <strong>{}</strong> ج.م<br>'
#             'مستحق: <strong>{}</strong> ج.م',
#             obj.archive_total_payments,
#             obj.archive_total_owed
#         )
#     archive_financial_summary.short_description = 'الملخص المالي'

#     def has_add_permission(self, request):
#         """منع الإضافة المباشرة - الأرشفة تتم تلقائياً"""
#         return False
